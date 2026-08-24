"""
build_liquidacion_month.py
==========================
Paso 1 del cierre mensual BIG.

Copia estructura del ultimo mes disponible (ej 'Mayo') como nueva hoja del mes
nuevo (ej 'Junio') en Liuidacion Global BIG 2026.xlsx. Pega los datos NAV/fees
del template ProCapital para el mes correspondiente, matcheando por fecha
(row 11) y por label (col A). Las formulas rows 41-49 (desglose PRO/PAMPA/
Premium/FA/FER) se copian con la hoja y se recalculan automaticamente.

Output: valor total PAMPA del mes (row 47 col AI o equivalente) que alimenta
el Paso 2.

Usage:
    python scripts/build_liquidacion_month.py --month 6 --year 2026
"""
import argparse
import shutil
from datetime import datetime, date, timedelta
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

FOLDER = Path("C:/Users/lmonp/Dropbox/Maximus BIG/2026/Luiquidacion Comisiones BIG")
DEFAULT_PROCAPITAL = FOLDER / "ProCapital_XS3037627794 (1).xlsx"
DEFAULT_LIQUIDACION = FOLDER / "Liuidación Global BIG 2026.xlsx"

# Formula de cada fila de comision para una columna de dia.
#   col  = la columna del dia que se esta escribiendo
#   prev = la columna anterior (para el primer dia, C = anchor del cierre del
#          mes anterior)
ROW_FORMULA = {
    42: lambda col, prev: f"={prev}38*$B$42/365",
    44: lambda col, prev: f"={col}41*$B$44/365",
    45: lambda col, prev: f"=({prev}38-{prev}41)*$B$45/365",
    46: lambda col, prev: f"={prev}38*$B$46/365",
    47: lambda col, prev: f"=({prev}38-{prev}41)*$B$47/365+1.3%*{prev}41/365",
    48: lambda col, prev: f"={prev}38*$B$48/365",
    49: lambda col, prev: f"={col}47-{col}48",
}

# Etiqueta esperada en la col A de cada fila que ROW_FORMULA escribe o
# referencia. ROW_FORMULA hardcodea numeros de fila: si alguien inserta o
# borra una fila en la plantilla, las formulas se escribirian en las filas
# equivocadas SIN fallar, y el Excel saldria silenciosamente mal. Este guard
# corta antes de tocar nada.
EXPECTED_LABELS = {
    38: "CAV",           # referenciada por 42/45/46/47/48
    42: "PRO + PAMPA",
    44: "PRO",
    45: "PRO",
    46: "Premium",
    47: "PAMPA",
    48: "FA",
    49: "FER",
}


def assert_layout(ws):
    """Verifica que el bloque de comisiones este donde ROW_FORMULA cree.

    Devuelve lista de problemas (vacia = todo OK).
    """
    problemas = []
    for row, esperado in EXPECTED_LABELS.items():
        actual = ws.cell(row=row, column=1).value
        actual = actual.strip() if isinstance(actual, str) else actual
        if actual != esperado:
            problemas.append(f"fila {row}: esperaba col A = '{esperado}', encontre {actual!r}")

    # Fila 41 (nominal de la nota) no tiene etiqueta pero SI la referencian las
    # formulas de 44/45/47: chequeamos que siga siendo la fila numerica.
    a41 = ws.cell(row=41, column=1).value
    b41 = ws.cell(row=41, column=2).value
    if a41 not in (None, ""):
        problemas.append(f"fila 41: esperaba col A vacia, encontre {a41!r}")
    if not isinstance(b41, (int, float)):
        problemas.append(f"fila 41: esperaba col B numerica (nominal nota), encontre {b41!r}")

    return problemas

MONTH_NAMES = {
    1: "ENE", 2: "FEB", 3: "MAR", 4: "Abril", 5: "Mayo",
    6: "Junio", 7: "Julio", 8: "Agosto", 9: "Septiembre",
    10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

# Mapeo (Liq label) -> (Proc label). Match por label en col A.
LIQ_TO_PROC_MAP = {
    "Notes Subscriptions": "Notes Subscriptions (+)",
    "Notes Redemptions": "Notes Redemptions (-)",
    "Total Nominal Notes": "Start. Nominal Notes",
    "Balance 1": "Cust. Balance 1",
    "Balance 2": "Cust. Balance 2",
    "Cash Subscriptions": "Cash Subscriptions (+)",
    "Cash Redemptions": "Cash Redemptions (-)",
    "Withdrawals": "Withdrawals",
    "Broker Charges": None,
    "Extraordinary Fees": "Extraordinary Fees",
    "Credit": "Total Credit",
    "Total Balance": "Total Balance",
    "EUR/USD Rate": "EUR/USD FX Rate",
    "Management Fee": "Management Fee",
    "Maintenance Fee": "Maintenance Fee",
    "Extraordinary Fee": "Extraordinary Fees",
    "A. M. F. in Inst.": "Annual Access Fee",
    "Set Up Fee in Inst.": "Setup Fee",
    "Total Fees": "Total Fees",
    "Accrued Fees": "Accrued Fees",
    "Nominal Notes": "Nominal Notes",
    "CAV": "CAV",
    "NAV": "NAV",
    "% Change": "% Change",
}


def find_row_by_label(ws, label, max_row=100):
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() == label:
            return r
    return None


def get_proc_month_cols(ws_proc, year, month):
    """Devuelve dict {col_index: date} para todas las cols del mes."""
    result = {}
    for c in range(3, ws_proc.max_column + 1):
        v = ws_proc.cell(row=11, column=c).value
        if v is not None and hasattr(v, 'year'):
            if v.year == year and v.month == month:
                result[c] = v.date() if hasattr(v, 'date') else v
    return result


def get_proc_anchor_col(ws_proc, year, month):
    """Devuelve col del ultimo dia del mes anterior (anchor)."""
    if month == 1:
        anchor = date(year - 1, 12, 31)
    else:
        anchor = date(year, month, 1) - timedelta(days=1)
    for c in range(3, ws_proc.max_column + 1):
        v = ws_proc.cell(row=11, column=c).value
        if v is not None and hasattr(v, 'year'):
            v_date = v.date() if hasattr(v, 'date') else v
            if v_date == anchor:
                return c, anchor
    return None, anchor


def backup_file(path: Path):
    backup_dir = path.parent / ".backups"
    backup_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = backup_dir / f"{path.stem}.{ts}{path.suffix}"
    shutil.copy2(path, backup_path)
    return backup_path


def days_in_month(year, month):
    if month == 12:
        next_first = date(year + 1, 1, 1)
    else:
        next_first = date(year, month + 1, 1)
    last = next_first - timedelta(days=1)
    return last.day


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", type=int, required=True, help="Mes (1-12)")
    ap.add_argument("--year", type=int, required=True)
    ap.add_argument("--procapital", type=str, default=str(DEFAULT_PROCAPITAL))
    ap.add_argument("--liquidacion", type=str, default=str(DEFAULT_LIQUIDACION))
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--force", action="store_true", help="Sobreescribe hoja si existe")
    args = ap.parse_args()

    proc_path = Path(args.procapital)
    liq_path = Path(args.liquidacion)
    year = args.year
    month = args.month
    month_name = MONTH_NAMES[month]
    n_days = days_in_month(year, month)

    if not proc_path.exists():
        raise FileNotFoundError(f"ProCapital no encontrado: {proc_path}")
    if not liq_path.exists():
        raise FileNotFoundError(f"Liquidacion Global no encontrado: {liq_path}")

    print(f"[{datetime.now().isoformat(timespec='seconds')}] Paso 1: Liquidacion Global {month_name} {year}")
    print(f"  Dias del mes: {n_days}")

    wb_proc = openpyxl.load_workbook(str(proc_path), data_only=True)
    ws_proc = wb_proc["NAV Calculation"]
    proc_anchor_col, anchor_date = get_proc_anchor_col(ws_proc, year, month)
    proc_month_cols = get_proc_month_cols(ws_proc, year, month)
    if not proc_month_cols:
        raise ValueError(f"ProCapital no tiene datos para {year}-{month:02d}")
    if len(proc_month_cols) < n_days:
        print(f"  WARN: ProCapital tiene {len(proc_month_cols)} cols del mes, esperaba {n_days}")
    anchor_str = get_column_letter(proc_anchor_col) if proc_anchor_col else 'N/A'
    print(f"  ProCapital anchor col: {anchor_str} ({anchor_date})")
    first_col = get_column_letter(min(proc_month_cols))
    last_col = get_column_letter(max(proc_month_cols))
    print(f"  ProCapital cols mes: {first_col}->{last_col} ({len(proc_month_cols)} dias)")

    wb_liq = openpyxl.load_workbook(str(liq_path))
    existing_sheets = wb_liq.sheetnames
    template_name = existing_sheets[-1]
    print(f"  Template Liq: '{template_name}'")

    problemas = assert_layout(wb_liq[template_name])
    if problemas:
        print("  ERROR: la plantilla no tiene el layout que este script espera.")
        print("         ROW_FORMULA hardcodea las filas 42/44-49 (y referencia 38/41).")
        print("         Si se movio alguna fila, hay que actualizar ROW_FORMULA +")
        print("         EXPECTED_LABELS antes de correr esto, si no el Excel sale mal.")
        for prob in problemas:
            print(f"         - {prob}")
        return
    print(f"  Layout OK: filas de comision 42/44-49 verificadas contra col A")

    if month_name in wb_liq.sheetnames:
        if args.force:
            del wb_liq[month_name]
        else:
            print(f"  ERROR: Hoja '{month_name}' ya existe. Usar --force para sobreescribir.")
            return

    if not args.dry_run:
        bpath = backup_file(liq_path)
        print(f"  Backup: {bpath.name}")

    ws_template = wb_liq[template_name]
    ws_new = wb_liq.copy_worksheet(ws_template)
    ws_new.title = month_name
    print(f"  Hoja '{month_name}' creada (copia de '{template_name}')")

    for c in range(3, ws_new.max_column + 1):
        ws_new.cell(row=11, column=c).value = None

    ws_new.cell(row=11, column=3).value = datetime(anchor_date.year, anchor_date.month, anchor_date.day)
    ws_new.cell(row=11, column=3).number_format = "m/d/yyyy"
    proc_cols_sorted = sorted(proc_month_cols.keys())
    for i, proc_col in enumerate(proc_cols_sorted):
        liq_col = 4 + i
        dt = proc_month_cols[proc_col]
        ws_new.cell(row=11, column=liq_col).value = datetime(dt.year, dt.month, dt.day)
        ws_new.cell(row=11, column=liq_col).number_format = "m/d/yyyy"

    last_data_col = 4 + len(proc_cols_sorted) - 1
    total_col = last_data_col + 1
    print(f"  Fila 11: C={anchor_date} + D->{get_column_letter(last_data_col)} ({len(proc_cols_sorted)} dias), SUM col={get_column_letter(total_col)}")

    copied = 0
    skipped = 0
    for liq_label, proc_label in LIQ_TO_PROC_MAP.items():
        liq_row = find_row_by_label(ws_new, liq_label, max_row=45)
        if liq_row is None:
            print(f"  WARN: Liq label no encontrado: '{liq_label}'")
            continue
        if proc_label is None:
            skipped += 1
            continue
        proc_row = find_row_by_label(ws_proc, proc_label, max_row=70)
        if proc_row is None:
            print(f"  WARN: Proc label no encontrado: '{proc_label}'")
            continue
        if proc_anchor_col:
            v = ws_proc.cell(row=proc_row, column=proc_anchor_col).value
            ws_new.cell(row=liq_row, column=3).value = v
        for i, proc_col in enumerate(proc_cols_sorted):
            liq_col = 4 + i
            v = ws_proc.cell(row=proc_row, column=proc_col).value
            ws_new.cell(row=liq_row, column=liq_col).value = v
        copied += 1

    print(f"  Rows copiadas: {copied} | Skipped: {skipped}")

    template_last_data_col = None
    for c in range(ws_new.max_column, 3, -1):
        v = ws_new.cell(row=47, column=c).value
        if isinstance(v, str) and v.startswith("=SUM("):
            template_last_data_col = c - 1
            break

    if template_last_data_col and template_last_data_col > last_data_col:
        rows_to_clear = [11, 13, 14, 15, 17, 18, 19, 20, 21, 22, 23, 24, 25, 27,
                         29, 30, 31, 32, 33, 34, 36, 37, 38, 39, 40, 42, 44, 45, 46, 47, 48, 49]
        for r in rows_to_clear:
            for c in range(last_data_col + 1, template_last_data_col + 2):
                ws_new.cell(row=r, column=c).value = None
        print(f"  Cols extras limpiadas: {get_column_letter(last_data_col+1)}->{get_column_letter(template_last_data_col+1)}")

    # BUGFIX 2026-08-24: cuando el mes nuevo tiene IGUAL O MAS dias que la
    # plantilla (ej Junio 30 dias -> Julio 31 dias), la columna que en la
    # plantilla tenia la formula "=SUM(...)" (su columna de total) pasa a ser
    # un dia real del mes nuevo. copy_worksheet() no la convierte -- queda con
    # la vieja formula SUM, y el total nuevo terminaba sumando esa columna
    # (que ya era una suma parcial) DOS VECES, mientras el ultimo dia real
    # nunca se calculaba. Detectado por Lucas en Julio 2026 (revisar
    # workflow_cierre_mensual_big.md). Fix: regenerar en codigo, para TODAS
    # las columnas de dia (D..last_data_col), la formula de cada fila, en vez
    # de confiar en lo que trajo copy_worksheet.
    for c in range(4, last_data_col + 1):
        col = get_column_letter(c)
        prev = get_column_letter(c - 1)
        for r, fn in ROW_FORMULA.items():
            ws_new.cell(row=r, column=c).value = fn(col, prev)

    last_col_letter = get_column_letter(last_data_col)
    for r in ROW_FORMULA:
        ws_new.cell(row=r, column=total_col).value = f"=SUM(D{r}:{last_col_letter}{r})"

    print(f"  Formulas SUM reescritas en col {get_column_letter(total_col)}, rows 42/44/45/46/47/48/49")

    if args.dry_run:
        print("  DRY RUN — no se guarda")
        return None
    wb_liq.save(str(liq_path))
    print(f"  Guardado: {liq_path.name}")

    wb_calc = openpyxl.load_workbook(str(liq_path), data_only=True)
    ws_calc = wb_calc[month_name]
    pampa_total = ws_calc.cell(row=47, column=total_col).value

    if pampa_total is None or not isinstance(pampa_total, (int, float)):
        # openpyxl no evaluo la formula. Replicar formula manualmente sobre CAV raw.
        # PAMPA daily = (CAV - 20M) * 0.0135 / 365 + 0.013 * 20M / 365
        threshold = ws_calc.cell(row=41, column=3).value or 20_000_000
        pampa_rate = ws_calc.cell(row=47, column=2).value or 0.0135
        below_rate = 0.013  # Rate para los primeros 20M
        pampa_total = 0
        for c in range(4, last_data_col + 1):
            cav = ws_calc.cell(row=38, column=c).value
            if cav is None or not isinstance(cav, (int, float)):
                continue
            pampa_total += (cav - threshold) * pampa_rate / 365 + below_rate * threshold / 365
        source = "calculado manualmente desde CAV row 38 (openpyxl no evalua formulas)"
    else:
        source = f"row 47 col {get_column_letter(total_col)}"

    print()
    if pampa_total:
        print(f"  === PAMPA TOTAL {month_name} {year} = ${pampa_total:,.2f} ===")
        print(f"  ({source})")
    else:
        print(f"  WARN: PAMPA total no pudo calcularse")

    return pampa_total


if __name__ == "__main__":
    main()
