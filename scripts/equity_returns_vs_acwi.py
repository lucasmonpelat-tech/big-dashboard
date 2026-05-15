"""
equity_returns_vs_acwi.py
=========================
"Buy-and-hold price race" — la carrera del activo desde la primera compra:

Para cada activo:
  Asset Return = (Precio HOY − Precio PRIMER buy) / Precio PRIMER buy
  ACWI Return  = (ACWI HOY − ACWI EN fecha del primer buy) / ACWI EN fecha del primer buy
  Alpha = Asset Return − ACWI Return

Para holdings CERRADOS: "HOY" se reemplaza por la fecha del último sell, y "precio HOY"
por el precio del último sell.

NO depende de DCA / size de los buys subsiguientes — mide la decisión de elegir el
activo en una ventana de tiempo dada, vs el benchmark en esa misma ventana.

Trabaja a TRADE-LEVEL (precios y fechas reales de Pershing).

Usage:
    python scripts/equity_returns_vs_acwi.py
"""

import openpyxl
import json
import re
from datetime import datetime, date
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).parent.parent

# Default location of the Transactions Excel
DEFAULT_XLSX = Path("C:/Users/lmonp/Dropbox/2026/Pampa BIG/Transactions_JXD101380 History Analisis Equity Race.xlsx")

# Security code → (sleeve, ticker, isin, yahoo_ticker_for_price)
# Solo equity holdings (FI/Alts se podrian agregar despues)
SECURITY_MAP = {
    "CSPX:GB":   ("Equity", "CSPX",  "IE00B5BMR087", "SPY"),
    "BRK B":     ("Equity", "BRK.B", "US0846707026", "BRK-B"),
    "G6430T809": ("Equity", "NBGMT", "IE00BFMHRK20", None),
    "L6366L196": ("Equity", "MFSCV", "LU1985812756", None),
    "L468AA656": ("Equity", "JHGSC", "LU2940405447", None),
    "G5S05E528": ("Equity", "LGLI",  "IE00BF4KN675", None),
    "G5441Y310": ("Equity", "NBRE",  "LU1648392275", None),
    "G9373W177": ("Equity", "VIRTUS","LU2020382479", None),
    "G8T49N214": ("Equity", "THOR",  "IE00B6YCBF59", None),
    "ILF":       ("Equity", "ILF",   "US4642873909", "ILF"),
    "4BRZ:DE":   ("Equity", "4BRZ",  "DE000A0Q4R85", "EWZ"),
    "ARGT":      ("Equity", "ARGT",  "US37950E2596", "ARGT"),
}

NAME_BY_TICKER = {
    "CSPX":   "iShares Core S&P 500 UCITS",
    "BRK.B":  "Berkshire Hathaway B (CLOSED Apr-2026)",
    "NBGMT":  "NB Global Equity Megatrends I",
    "MFSCV":  "MFS Meridian Contrarian Value I1",
    "JHGSC":  "Janus Henderson Global Smaller Cos F2",
    "LGLI":   "Lazard Global Listed Infrastructure A",
    "NBRE":   "NB US Real Estate (CLOSED Feb-2026)",
    "VIRTUS": "Virtus US Small Cap (CLOSED Feb-2026)",
    "THOR":   "Thornburg Equity Income Builder I",
    "ILF":    "iShares Latin America 40 ETF",
    "4BRZ":   "iShares MSCI Brazil UCITS (DE)",
    "ARGT":   "Global X MSCI Argentina ETF",
}


def parse_transactions(xlsx_path: Path):
    """Read Pershing Transactions export. Returns list of trade dicts."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["ExportExcel"] if "ExportExcel" in wb.sheetnames else wb.active

    # Find header row
    hdr_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[0] == "Process Date":
            hdr_row = i
            break
    if hdr_row is None:
        raise ValueError("Could not find header row 'Process Date'")

    trades = []
    patterns = [
        re.compile(r"^(Buy|Sell|Correct Buy|Correct Sell)\s+(-?[\d.]+)\s+(?:share\(s\)|parValue)\s+of\s+([A-Z0-9: .]+?)\s+at\s+([\d.]+)$"),
        re.compile(r"^(Buy|Sell|Correct Buy|Correct Sell)\s+(-?[\d.]+)\s+of\s+([A-Z0-9: .]+?)\s+at\s+([\d.]+)$"),
    ]

    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        date_cell, typ, desc, net_base, *_ = row
        if not date_cell or not typ:
            continue
        typ_s = str(typ).strip()

        if not any(k in typ_s for k in ["Buy", "Sell"]):
            continue
        if typ_s.startswith("Cancel"):
            continue

        action = None
        qty = None
        sym = None
        price = None
        for pat in patterns:
            m = pat.match(typ_s)
            if m:
                action, qty, sym, price = m.groups()
                break
        if not action:
            continue

        trade_date = date_cell if isinstance(date_cell, datetime) else None
        if not trade_date:
            continue

        side = "SELL" if "Sell" in action else "BUY"
        qty_val = abs(float(qty))
        net_usd = float(net_base) if net_base else 0

        trades.append({
            "date": trade_date.date(),
            "side": side,
            "sym": sym.strip(),
            "qty": qty_val,
            "price": float(price),
            "net_usd": net_usd,
            "raw_type": typ_s,
        })

    return trades


def aggregate_by_ticker(trades):
    """Group trades by ticker. Returns dict[ticker] -> aggregated metrics."""
    by_sym = defaultdict(list)
    for t in trades:
        meta = SECURITY_MAP.get(t["sym"])
        if not meta:
            continue
        if meta[0] != "Equity":
            continue
        by_sym[meta[1]].append(t)  # ticker

    aggregates = {}
    for ticker, ts in by_sym.items():
        ts_sorted = sorted(ts, key=lambda x: x["date"])
        buys = [t for t in ts_sorted if t["side"] == "BUY"]
        sells = [t for t in ts_sorted if t["side"] == "SELL"]

        total_buy_usd = sum(-t["net_usd"] for t in buys)  # net_usd is negative for buys; flip sign
        total_sell_usd = sum(t["net_usd"] for t in sells)
        qty_bought = sum(t["qty"] for t in buys)
        qty_sold = sum(t["qty"] for t in sells)
        qty_remaining = qty_bought - qty_sold

        # Anchor para buy-and-hold race: primer buy que sea >=25% del total.
        # Skipea "test buys" chicos (ej: $10K test antes de una big buy de $840K).
        # Si ningun buy individual cruza el 25%, fallback al primer buy cronologico.
        ANCHOR_THRESHOLD_PCT = 25.0
        first_buy_date = None
        first_buy_price = None
        if buys:
            threshold_amt = total_buy_usd * ANCHOR_THRESHOLD_PCT / 100
            for t in buys:  # buys ya esta ordenado por fecha
                amt = t["qty"] * t["price"]
                if amt >= threshold_amt:
                    first_buy_date = t["date"]
                    first_buy_price = t["price"]
                    break
            # Fallback: si ningun buy llega al 25%, usar el primero cronologico
            if first_buy_date is None:
                first_buy_date = buys[0]["date"]
                first_buy_price = buys[0]["price"]

        last_sell_date = sells[-1]["date"] if sells else None
        last_sell_price = sells[-1]["price"] if sells else None
        last_trade_date = ts_sorted[-1]["date"]

        aggregates[ticker] = {
            "ticker": ticker,
            "trades": ts_sorted,
            "total_buy_usd": total_buy_usd,
            "total_sell_usd": total_sell_usd,
            "qty_bought": qty_bought,
            "qty_sold": qty_sold,
            "qty_remaining": qty_remaining,
            "first_buy_date": first_buy_date,
            "first_buy_price": first_buy_price,
            "last_sell_date": last_sell_date,
            "last_sell_price": last_sell_price,
            "last_trade_date": last_trade_date,
            "is_closed": qty_remaining < 0.01,
            "avg_buy_price": (total_buy_usd / qty_bought) if qty_bought > 0 else None,
            "avg_sell_price": (total_sell_usd / qty_sold) if qty_sold > 0 else None,
            "n_buys": len(buys),
            "n_sells": len(sells),
        }
    return aggregates


def fetch_acwi_prices(date_list):
    """Fetch ACWI close prices for a list of dates. Returns dict[date] -> price."""
    try:
        import yfinance as yf
    except ImportError:
        print("ERROR: yfinance not installed. Run: pip install yfinance")
        return {}

    if not date_list:
        return {}

    min_d = min(date_list)
    max_d = max(date_list)
    # Get a wider range to ensure trading days are captured
    from datetime import timedelta
    start = min_d - timedelta(days=10)
    end = max_d + timedelta(days=10)

    print(f"  Fetching ACWI from {start} to {end}...")
    ticker = yf.Ticker("ACWI")
    hist = ticker.history(start=start.isoformat(), end=end.isoformat())
    if hist.empty:
        return {}

    # Build a date->price lookup; if target date is not a trading day, use the closest prior trading day
    prices = {d.date(): float(row["Close"]) for d, row in hist.iterrows()}
    return prices


def acwi_price_at_or_before(prices, target_date):
    """Get ACWI price at target_date or the most recent prior trading day."""
    from datetime import timedelta
    for delta in range(0, 15):
        check_d = target_date - timedelta(days=delta)
        if check_d in prices:
            return prices[check_d]
    return None


def get_current_mv(ticker):
    """Get current MV from positions_latest.json (for still-held positions)."""
    pos_file = ROOT / "data" / "positions_latest.json"
    if not pos_file.exists():
        return None
    with open(pos_file) as f:
        d = json.load(f)
    for p in d["positions"]:
        if p.get("ticker") == ticker:
            return p.get("value")
    return None


def compute_returns(aggregates, acwi_prices, calculation_date=None):
    """Buy-and-hold price race: para cada activo, compara el retorno del PRECIO
    desde primer buy hasta hoy (o last sell si esta cerrado) contra ACWI en la
    misma ventana.

    Asset Return = (end_price - first_buy_price) / first_buy_price
    ACWI Return  = (acwi_end - acwi_start) / acwi_start
    Alpha = Asset - ACWI

    NO depende de DCA / size de buys subsiguientes — mide la decision de elegir
    el activo en una ventana de tiempo dada.
    """
    calc_date = calculation_date or date.today()
    results = []

    for ticker, agg in aggregates.items():
        first_buy_date = agg["first_buy_date"]
        first_buy_price = agg["first_buy_price"]
        is_closed = agg["is_closed"]

        # End date + end price segun open/closed
        if is_closed:
            end_date = agg["last_sell_date"]
            end_price = agg["last_sell_price"]
            current_mv = 0
        else:
            end_date = calc_date
            current_mv = get_current_mv(ticker) or 0
            qty_rem = agg["qty_remaining"]
            # current_price = MV / qty_remaining (works for ETFs y UCITS uniformemente)
            end_price = (current_mv / qty_rem) if qty_rem > 0 else None

        # ===== Asset return (buy-and-hold price race) =====
        if first_buy_price and end_price and first_buy_price > 0:
            return_pct = (end_price - first_buy_price) / first_buy_price * 100
        else:
            return_pct = None

        # ===== ACWI return en la misma ventana =====
        acwi_start = acwi_price_at_or_before(acwi_prices, first_buy_date)
        acwi_end = acwi_price_at_or_before(acwi_prices, end_date)
        if acwi_start and acwi_end and acwi_start > 0:
            acwi_return_pct = (acwi_end / acwi_start - 1) * 100
        else:
            acwi_return_pct = None

        alpha_pct = (return_pct - acwi_return_pct) if (return_pct is not None and acwi_return_pct is not None) else None

        # Total proceeds para info (lo que entro como cash, no es input del return)
        total_proceeds = agg["total_sell_usd"] + current_mv

        results.append({
            "ticker": ticker,
            "name": NAME_BY_TICKER.get(ticker, ticker),
            "is_closed": is_closed,
            "first_buy_date": first_buy_date.isoformat() if first_buy_date else None,
            "first_buy_price": round(first_buy_price, 4) if first_buy_price else None,
            "end_date": end_date.isoformat() if end_date else None,
            "end_price": round(end_price, 4) if end_price else None,
            "n_buys": agg["n_buys"],
            "n_sells": agg["n_sells"],
            "qty_bought": round(agg["qty_bought"], 2),
            "qty_sold": round(agg["qty_sold"], 2),
            "qty_remaining": round(agg["qty_remaining"], 2),
            "total_buy_usd": round(agg["total_buy_usd"], 2),
            "total_sell_usd": round(agg["total_sell_usd"], 2),
            "current_mv_usd": round(current_mv, 2) if not is_closed else 0,
            "total_proceeds_usd": round(total_proceeds, 2),
            "avg_buy_price": round(agg["avg_buy_price"], 4) if agg["avg_buy_price"] else None,
            "avg_sell_price": round(agg["avg_sell_price"], 4) if agg["avg_sell_price"] else None,
            "return_pct": round(return_pct, 2) if return_pct is not None else None,
            "acwi_start": round(acwi_start, 2) if acwi_start else None,
            "acwi_end": round(acwi_end, 2) if acwi_end else None,
            "acwi_return_pct": round(acwi_return_pct, 2) if acwi_return_pct is not None else None,
            "alpha_pct": round(alpha_pct, 2) if alpha_pct is not None else None,
        })

    # Sort by alpha descending (best first)
    results.sort(key=lambda x: -(x["alpha_pct"] if x["alpha_pct"] is not None else -999))
    return results


def status_label(alpha, is_closed):
    if alpha is None:
        return ("unknown", "[?] UNKNOWN")
    if alpha > 0:
        return ("outperform" + ("_closed" if is_closed else ""), "[+] OUTPERFORM" + (" (cerrado)" if is_closed else ""))
    elif alpha < 0:
        return ("underperform" + ("_closed" if is_closed else ""), "[-] UNDERPERFORM" + (" (cerrado)" if is_closed else ""))
    else:
        return ("neutral" + ("_closed" if is_closed else ""), "[=] NEUTRAL")


def main():
    import sys
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        print(f"ERROR: {xlsx} no existe")
        return 1
    print(f"Reading: {xlsx.name}")
    trades = parse_transactions(xlsx)
    print(f"Parsed {len(trades)} trades")

    aggregates = aggregate_by_ticker(trades)
    print(f"Equity tickers: {len(aggregates)}")

    # Determine calculation_date from positions_latest.json (so MV and ACWI match)
    pos_file = ROOT / "data" / "positions_latest.json"
    calc_date = date.today()
    if pos_file.exists():
        with open(pos_file) as f:
            pos = json.load(f)
        as_of_str = pos.get("as_of", "")
        # Parse "May 5, 2026 9:35 AM EDT"
        try:
            calc_date = datetime.strptime(as_of_str.split(" 9")[0].split(" ")[0] + " " + as_of_str.split(" ")[1].rstrip(',') + " " + as_of_str.split(" ")[2], "%b %d %Y").date()
        except Exception:
            try:
                # Try alternative parse
                import re as _re
                m = _re.search(r"(\w+)\s+(\d+),\s+(\d+)", as_of_str)
                if m:
                    months = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
                    calc_date = date(int(m.group(3)), months[m.group(1)], int(m.group(2)))
            except Exception:
                pass
    print(f"Calculation date: {calc_date} (matches MV and ACWI end)")

    # Collect all relevant dates for ACWI fetch
    all_dates = set()
    for agg in aggregates.values():
        if agg["first_buy_date"]:
            all_dates.add(agg["first_buy_date"])
        if agg["last_trade_date"]:
            all_dates.add(agg["last_trade_date"])
    all_dates.add(calc_date)

    acwi_prices = fetch_acwi_prices(list(all_dates))
    print(f"ACWI prices fetched: {len(acwi_prices)} days")

    results = compute_returns(aggregates, acwi_prices, calculation_date=calc_date)

    # Add status
    for r in results:
        s_code, s_label = status_label(r["alpha_pct"], r["is_closed"])
        r["status"] = s_code
        r["status_label"] = s_label

    # Print report
    print(f"\n{'=' * 140}")
    print(f"EQUITY HOLDINGS — RETURN vs ACWI (Lucas methodology — transaction-level)")
    print(f"{'=' * 140}")
    print(f"{'Status':30s} {'Ticker':8s} {'First Buy':12s} {'End Date':12s} "
          f"{'Bought $':>13s} {'Sold $':>13s} {'MV Now':>11s} "
          f"{'Return%':>8s} {'ACWI%':>8s} {'Alpha%':>9s}")
    print("-" * 140)
    for r in results:
        mv_str = f"${r['current_mv_usd']:>9,.0f}" if not r["is_closed"] else "    (closed)"
        ret_str = f"{r['return_pct']:>+7.2f}%" if r['return_pct'] is not None else "    —  "
        acwi_str = f"{r['acwi_return_pct']:>+7.2f}%" if r['acwi_return_pct'] is not None else "    —  "
        alpha_str = f"{r['alpha_pct']:>+8.2f}%" if r['alpha_pct'] is not None else "     —  "
        print(
            f"{r['status_label']:30s} "
            f"{r['ticker']:8s} "
            f"{r['first_buy_date']:12s} "
            f"{r['end_date']:12s} "
            f"${r['total_buy_usd']:>11,.0f} "
            f"${r['total_sell_usd']:>11,.0f} "
            f"{mv_str:>11s} "
            f"{ret_str} "
            f"{acwi_str} "
            f"{alpha_str}"
        )

    # Output 1: detailed file (this script's output)
    out = {
        "refreshedAt": datetime.now().isoformat(),
        "source": f"Pershing transactions (trade-level) — {xlsx.name}",
        "method": "Buy-and-hold price race: Asset Return = (end_price - first_buy_price) / first_buy_price. ACWI Return = same window. Alpha = Asset - ACWI.",
        "holdings": results,
    }
    out_path = ROOT / "data" / "equity_returns_vs_acwi.json"
    with open(out_path, "w") as f:
        json.dump(out, f, indent=2)
    print(f"\n[OK] Saved detailed: {out_path}")

    # Output 2: dashboard-compatible (overwrites old Mod Dietz output)
    def months_between(d1_iso, d2_iso):
        if not d1_iso or not d2_iso:
            return None
        d1 = datetime.fromisoformat(d1_iso).date() if isinstance(d1_iso, str) else d1_iso
        d2 = datetime.fromisoformat(d2_iso).date() if isinstance(d2_iso, str) else d2_iso
        return (d2.year - d1.year) * 12 + (d2.month - d1.month)

    dashboard_holdings = []
    for r in results:
        dashboard_holdings.append({
            "ticker": r["ticker"],
            "name": r["name"],
            "period_start": r["first_buy_date"],
            "period_end": r["end_date"],
            "months_held": months_between(r["first_buy_date"], r["end_date"]),
            "twr_pct": None,  # Not computed in this methodology
            "mwr_pct": r["return_pct"],
            "net_pnl_usd": round(r["total_proceeds_usd"] - r["total_buy_usd"], 0) if r["return_pct"] is not None else None,
            "cash_in_usd": r["total_buy_usd"],
            "cash_out_usd": r["total_sell_usd"] + (r["current_mv_usd"] if not r["is_closed"] else 0),
            "acwi_period_return_pct": r["acwi_return_pct"],
            "alpha_twr_pp": None,
            "alpha_mwr_pp": r["alpha_pct"],
            "avg_weight_pct": None,
            "contribution_pp": None,
            "is_closed": r["is_closed"],
            "status": r["status"],
            "status_label": r["status_label"],
            "final_mv_usd": r["current_mv_usd"] if not r["is_closed"] else 0,
        })

    dashboard_out = {
        "refreshedAt": datetime.now().isoformat(),
        "source": "Buy-and-hold price race vs ACWI (transactions Pershing)",
        "method": "Asset Return = (end_price - first_buy_price) / first_buy_price. ACWI Return = same window. Alpha = Asset - ACWI. End = today if open / last sell if closed.",
        "period_start": min(r["first_buy_date"] for r in results if r["first_buy_date"]),
        "period_end": max(r["end_date"] for r in results if r["end_date"]),
        "total_contribution_pp": None,
        "holdings": dashboard_holdings,
    }
    dashboard_path = ROOT / "data" / "equity_contributions_real.json"
    with open(dashboard_path, "w") as f:
        json.dump(dashboard_out, f, indent=2)
    print(f"[OK] Saved dashboard-compat: {dashboard_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
