"""
compute_fi_subasset.py
======================
Calcula el breakdown ponderado de sub-asset class del sleeve Fixed Income de BIG.

Entradas:
  - data/fi_subasset_lookup.json (breakdown por fondo, mantenido a mano)
  - Pershing Excel BIG_Posiciones_<date>_para_Maximus.xlsx (pesos $)

Logica:
  1. Lee el lookup.
  2. Lee el Pershing Excel y filtra el sleeve "Renta Fija" (sleeve mapping via
     ISIN_TO_SLEEVE del pershing_parser).
  3. Calcula pesos relativos (cada fondo / total FI).
  4. Multiplica weight x subasset_pct y suma por categoria.
  5. Valida que la suma sea ~100% (warning si <97 o >103).
  6. Escribe data/breakdowns/fi_subasset_big.json.

Uso:
    python scripts/compute_fi_subasset.py \
        --pershing-xlsx "C:/Users/lmonp/Downloads/BIG_Posiciones_31May2026_para_Maximus.xlsx"

Si no se pasa --pershing-xlsx, busca el ultimo BIG_Posiciones_*_para_Maximus.xlsx
en C:/Users/lmonp/Downloads/.
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

# Categorias canonicas (orden fijo para el output)
CATEGORIES = [
    "US Treasury",
    "Govt-related",
    "Govt Agency MBS",
    "Securitized Non-Agency",
    "Corporate IG",
    "Corporate HY",
    "EM Local + Hard",
    "Cat Bonds",
    "Bank Loans",
    "Other",
]

# Sleeves del Pershing parser — replicamos aca para no depender del import
# (pershing_parser usa ISIN_TO_SLEEVE; los 7 fondos FI estan abajo)
FI_ISINS = {
    "IE00BDT57R20",  # PIMCO LD
    "IE00B87KCF77",  # PIMCO Income
    "IE000OE87WX6",  # Man IG
    "IE00B29K0P99",  # PIMCO EM Local
    "XS2324777171",  # Tenac
    "LU2049315265",  # Schroder Cat Bond
    "IE00089T5MA6",  # Man EM Corp Credit
}


def find_latest_pershing(downloads_dir: Path) -> Path:
    candidates = sorted(
        downloads_dir.glob("BIG_Posiciones_*_para_Maximus.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(
            f"No se encontro ningun BIG_Posiciones_*_para_Maximus.xlsx en {downloads_dir}"
        )
    return candidates[0]


def read_pershing_fi_weights(xlsx_path: Path) -> dict[str, float]:
    """Devuelve {ISIN: valor_mercado_usd} solo para los 7 fondos FI.

    Soporta dos formatos:
      A) Pershing raw (header 'Description'):
         desc | cusip | MV | % | price | qty | ISIN | maturity | priceDate
      B) Maximus reformateado (header 'Sleeve'):
         Sleeve | Ticker | Activo | ISIN | Cantidad | Precio | Valor de Mercado USD | % AUM
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # Localizar fila header (Description = formato A, Sleeve = formato B)
    header_row = None
    fmt = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row[0] == "Description":
            header_row = i
            fmt = "pershing_raw"
            break
        if row[0] == "Sleeve":
            header_row = i
            fmt = "maximus"
            break
    if header_row is None:
        raise ValueError(f"No se encontro header (Description/Sleeve) en {xlsx_path}")

    weights: dict[str, float] = {}

    if fmt == "maximus":
        # Sleeve | Ticker | Activo | ISIN | Cantidad | Precio | MV | %AUM
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            sleeve = row[0]
            # Subtotales tienen sleeve=None y "Subtotal X" en col 2
            if sleeve is None:
                continue
            if str(sleeve).startswith("Subtotal"):
                continue
            isin = row[3]
            mv = row[6]  # columna 7 (1-indexed) = "Valor de Mercado USD"
            if mv is None or not isin:
                continue
            # Filtrar al sleeve FI (en este Excel se llama "Renta Fija")
            # pero seguimos validando contra FI_ISINS por las dudas
            if isin in FI_ISINS:
                weights[isin] = float(mv)
    else:
        # Detectar formato Pershing raw (new vs old)
        header_vals = [c for c in ws[header_row] if c.value is not None]
        new_format = any(str(c.value) == "ISIN" for c in header_vals[:8])
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if new_format:
                desc, _cusip, mv, _pct, _price, _qty, isin, *_ = row
            else:
                desc, _cusip, mv, _pct, _price, _qty, _sym, _u, _pa, isin, *_ = row
            if not desc or desc == "Disclaimer":
                break
            if mv is None or not isin:
                continue
            if isin in FI_ISINS:
                weights[isin] = float(mv)

    return weights


def compute_breakdown(
    lookup: dict, fund_weights_usd: dict[str, float]
) -> tuple[dict[str, float], dict[str, float], float]:
    """
    Devuelve (subasset_breakdown_pct, fund_weights_pct, fi_total_usd).

    subasset_breakdown_pct[categoria] = suma_i ( weight_i_pct * subasset_i_pct ) / 100
        (resultado en puntos porcentuales del sleeve FI)
    """
    fi_total = sum(fund_weights_usd.values())
    if fi_total <= 0:
        raise ValueError("Total FI = 0 — no hay posiciones FI en el Pershing")

    # pesos relativos en %
    weights_pct = {isin: (mv / fi_total) * 100 for isin, mv in fund_weights_usd.items()}

    breakdown = {cat: 0.0 for cat in CATEGORIES}
    funds_lookup = lookup["funds"]

    for isin, w_pct in weights_pct.items():
        fund = funds_lookup.get(isin)
        if not fund:
            print(f"[WARN] ISIN {isin} no esta en fi_subasset_lookup.json — se ignora")
            continue
        for cat in CATEGORIES:
            sub_pct = fund["subasset"].get(cat, 0)
            # weight (% del FI) * subasset (% del fondo) / 100 = puntos del FI
            breakdown[cat] += (w_pct * sub_pct) / 100

    # redondear para output
    breakdown = {cat: round(v, 2) for cat, v in breakdown.items()}
    weights_pct = {isin: round(v, 2) for isin, v in weights_pct.items()}

    return breakdown, weights_pct, fi_total


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--pershing-xlsx",
        type=Path,
        default=None,
        help="Path al Pershing Excel. Si se omite, busca el mas reciente en Downloads.",
    )
    parser.add_argument(
        "--lookup",
        type=Path,
        default=Path(__file__).parent.parent / "data" / "fi_subasset_lookup.json",
        help="Path al lookup JSON (default: data/fi_subasset_lookup.json).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).parent.parent / "data" / "breakdowns" / "fi_subasset_big.json",
        help="Path al output JSON (default: data/breakdowns/fi_subasset_big.json).",
    )
    args = parser.parse_args()

    # 1. Resolver Pershing path
    if args.pershing_xlsx is None:
        downloads = Path.home() / "Downloads"
        args.pershing_xlsx = find_latest_pershing(downloads)
        print(f"[INFO] Pershing auto-detectado: {args.pershing_xlsx}")
    else:
        if not args.pershing_xlsx.exists():
            print(f"[FAIL] Pershing no existe: {args.pershing_xlsx}")
            return 1

    # 2. Leer lookup
    if not args.lookup.exists():
        print(f"[FAIL] Lookup no existe: {args.lookup}")
        return 1
    with open(args.lookup, "r", encoding="utf-8") as f:
        lookup = json.load(f)
    print(f"[OK] Lookup cargado: {len(lookup['funds'])} fondos")

    # 3. Leer pesos del Pershing
    fund_weights_usd = read_pershing_fi_weights(args.pershing_xlsx)
    print(f"[OK] Pershing parseado: {len(fund_weights_usd)} fondos FI")
    for isin, mv in fund_weights_usd.items():
        name = lookup["funds"].get(isin, {}).get("name", "(no en lookup)")
        print(f"      {isin}  ${mv:>14,.2f}  {name}")

    # Detectar fondos FI en el Pershing que NO esten en el lookup
    missing = set(fund_weights_usd) - set(lookup["funds"])
    if missing:
        for isin in missing:
            print(f"[WARN] ISIN {isin} esta en Pershing FI pero NO en lookup — se ignora en breakdown")

    # 4. Calcular breakdown
    breakdown, weights_pct, fi_total = compute_breakdown(lookup, fund_weights_usd)
    total_sum = round(sum(breakdown.values()), 2)
    ok = 97.0 <= total_sum <= 103.0

    # 5. Output
    args.output.parent.mkdir(parents=True, exist_ok=True)
    out = {
        "computed_at": datetime.now().isoformat(timespec="seconds"),
        "source": "Weighted from fi_subasset_lookup.json x Pershing weights",
        "pershing_file": str(args.pershing_xlsx),
        "fi_total_aum_usd": round(fi_total, 2),
        "fund_weights": weights_pct,
        "subasset_breakdown_big": breakdown,
        "validation": {"sum": total_sum, "ok": ok},
    }
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)

    # 6. Report
    print()
    print("=" * 60)
    print(f"FI Total AUM:  ${fi_total:,.2f}")
    print("Sub-asset breakdown (% del sleeve FI):")
    for cat in CATEGORIES:
        print(f"  {cat:<25s} {breakdown[cat]:>6.2f}%")
    print(f"  {'TOTAL':<25s} {total_sum:>6.2f}%")
    print("=" * 60)

    if not ok:
        print(f"[WARN] Suma {total_sum}% fuera de rango [97, 103] — revisar lookup")
    else:
        print(f"[OK] Validacion: suma = {total_sum}% (dentro de [97, 103])")
    print(f"[OK] Output escrito en: {args.output}")

    return 0 if ok else 2


if __name__ == "__main__":
    sys.exit(main())
