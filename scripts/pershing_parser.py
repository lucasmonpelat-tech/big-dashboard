"""
pershing_parser.py
==================
Parses a Pershing 'Positions_JXDxxxxx.xlsx' export and generates the
updated BIG_POSITIONS array for data/funds_metadata.js.

Usage:
    python scripts/pershing_parser.py "path/to/Positions_JXD101380.xlsx"

Outputs:
    - Prints the new BIG_POSITIONS JavaScript array
    - Writes to data/positions_latest.json for traceability

External alts (NOT in Pershing) must be maintained manually via EXTERNAL_ALTS dict below.
"""

import openpyxl
import json
import sys
from datetime import datetime, date
from pathlib import Path

# External alternatives that don't appear in Pershing custodian
# (user invests in these separately, manually tracked)
EXTERNAL_ALTS = [
    {
        "isin": "LU2837777825",
        "ticker": "CALP",
        "name": "Carlyle AlpInvest Private Markets",
        "sleeve": "Alternatives",
        "value": 2722180.00,
        "ter": 1.00,
        "source": "Manual (last reported by Carlyle)"
    },
    {
        "isin": "FLEX-LEX",
        "ticker": "FLEX",
        "name": "Flex-Lexington Partners Secondaries (in transit)",
        "sleeve": "Alternatives",
        "value": 500000.00,  # Cash debited, position NAV pending first valuation
        "ter": None,
        "source": "Cash debited — NAV pending (secondaries fund)",
        "status": "IN_TRANSIT"
    },
    # Hamilton Lane Infrastructure — NOT YET funded (queued for Jun-26 cycle)
    # Do NOT include until cash actually debits.
]

# Map Pershing description patterns → standardized metadata
# Use ISIN as primary key, fall back on CUSIP for those without ISIN
ISIN_TO_SLEEVE = {
    # Equity
    "IE00B5BMR087": {"sleeve": "Equity", "ticker": "CSPX"},
    "IE00B6YCBF59": {"sleeve": "Equity", "ticker": "THOR"},
    "IE00BFMHRK20": {"sleeve": "Equity", "ticker": "NBGMT"},
    "LU1985812756": {"sleeve": "Equity", "ticker": "MFSCV"},
    "LU2940405447": {"sleeve": "Equity", "ticker": "JHGSC"},
    "IE00BF4KN675": {"sleeve": "Equity", "ticker": "LGLI"},
    "DE000A0Q4R85": {"sleeve": "Equity", "ticker": "4BRZ"},
    "US4642873909": {"sleeve": "Equity", "ticker": "ILF"},
    "US37950E2596": {"sleeve": "Equity", "ticker": "ARGT"},

    # Alternatives
    "US78463V1070": {"sleeve": "Alternatives", "ticker": "GLD"},
    "US46438F1012": {"sleeve": "Alternatives", "ticker": "IBIT"},
    "LU2659193242": {"sleeve": "Alternatives", "ticker": "NBPEA"},

    # Fixed Income
    "IE00BDT57R20": {"sleeve": "Fixed Income", "ticker": "PIMCO-LD"},
    "IE00B87KCF77": {"sleeve": "Fixed Income", "ticker": "PIMCO-INC"},
    "IE000OE87WX6": {"sleeve": "Fixed Income", "ticker": "MANIG"},
    "IE00B29K0P99": {"sleeve": "Fixed Income", "ticker": "PIMCO-EM"},
    "XS2324777171": {"sleeve": "Fixed Income", "ticker": "TGF"},
    "LU2049315265": {"sleeve": "Fixed Income", "ticker": "SGCB"},
}

CUSIP_TO_META = {
    "449LP0055": {"sleeve": "Alternatives", "ticker": "HLEND", "isin": "KYG4737U1085"},
    "379LP0083": {"sleeve": "Alternatives", "ticker": "GCRED", "isin": "GCRED-I"},
    "G7151GAA7": {"sleeve": "Alternatives", "ticker": "BPCC", "isin": "XS2658535526"},
    "USD999997": {"sleeve": "Cash", "ticker": "CASH", "isin": "CASH-USD"},
}

# TER data (kept from factsheets - not in Pershing)
TER_DATA = {
    "IE00B5BMR087": {"inst": 0.07, "a": None},
    "IE00B6YCBF59": {"inst": 0.89, "a": None},
    "IE00BFMHRK20": {"inst": 0.75, "a": 1.45},
    "LU1985812756": {"inst": 0.85, "a": 1.94},
    "LU2940405447": {"inst": 1.00, "a": None},
    "IE00BF4KN675": {"inst": 0.74, "a": None},
    "DE000A0Q4R85": {"inst": 0.47, "a": None},
    "US4642873909": {"inst": 0.59, "a": None},
    "US37950E2596": {"inst": 0.59, "a": None},
    "US78463V1070": {"inst": 0.25, "a": None},
    "US46438F1012": {"inst": 1.25, "a": None},
    "LU2659193242": {"inst": 0.40, "a": None},
    "IE00BDT57R20": {"inst": 0.55, "a": 1.45},
    "IE00B87KCF77": {"inst": 0.55, "a": 1.10},
    "IE000OE87WX6": {"inst": 0.89, "a": 1.89},
    "IE00B29K0P99": {"inst": 0.89, "a": None},
    "XS2324777171": {"inst": 0.75, "a": 1.41},
    "LU2049315265": {"inst": 1.37, "a": None},
    "KYG4737U1085": {"inst": 0.75, "a": None},
    "GCRED-I":      {"inst": 1.25, "a": None},
    "XS2658535526": {"inst": 1.25, "a": None},
    "CASH-USD":     {"inst": None, "a": None},
}

SHORT_NAME = {
    "IE00B5BMR087": "iShares Core S&P 500 UCITS",
    "IE00B6YCBF59": "Thornburg Equity Income Builder I",
    "IE00BFMHRK20": "NB Global Equity Megatrends I",
    "LU1985812756": "MFS Meridian Contrarian Value I1",
    "LU2940405447": "Janus Henderson Global Smaller Cos F2",
    "IE00BF4KN675": "Lazard Global Listed Infrastructure A",
    "DE000A0Q4R85": "iShares MSCI Brazil UCITS (DE)",
    "US4642873909": "iShares Latin America 40 ETF",
    "US37950E2596": "Global X MSCI Argentina ETF",
    "US78463V1070": "SPDR Gold Shares",
    "US46438F1012": "iShares Bitcoin Trust",
    "LU2659193242": "NB Global Private Equity Access Fund LI",
    "IE00BDT57R20": "PIMCO GIS Low Duration Income I",
    "IE00B87KCF77": "PIMCO GIS Income I",
    "IE000OE87WX6": "Man GLG Global IG Opportunities",
    "IE00B29K0P99": "PIMCO GIS EM Local Bond I",
    "XS2324777171": "Tenac Global Fund (TGF)",
    "LU2049315265": "Schroder GAIA Cat Bond Class C",
    "KYG4737U1085": "HPS Corporate Lending Fund",
    "GCRED-I":      "Golub Capital Private Credit",
    "XS2658535526": "Barings Private Credit Corporation (BPCC)",
    "CASH-USD":     "Cash USD",
}


def parse_pershing(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # Find header row
    as_of = None
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row[0] and str(row[0]).startswith("As of:"):
            as_of = str(row[0]).replace("As of:", "").strip()
        if row[0] == "Description":
            header_row = i
            break

    if header_row is None:
        raise ValueError("Could not find header row in Pershing export")

    # Detect column layout (Pershing has changed format over time)
    # New format (May 2026+): Description | SecID | MV | % | Price | Qty | ISIN | Maturity | PriceDate  (9 cols)
    # Old format:             Description | CUSIP | MV | % | Price | Qty | Symbol | _ | %Asset | ISIN  (10+ cols)
    header_vals = [c for c in ws[header_row] if c.value is not None]
    new_format = any("ISIN" == str(c.value) for c in header_vals[:8])  # ISIN appears in first 8 cols
    print(f"Detected {'NEW (9-col)' if new_format else 'OLD (10-col)'} Pershing format")

    # Parse positions
    positions = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if new_format:
            desc, cusip, mv, pct, price, qty, isin, maturity, price_date, *_ = row
        else:
            desc, cusip, mv, pct, price, qty, symbol, _u, pct_asset, isin, *_ = row
        if not desc or desc == "Disclaimer":
            break
        if mv is None:
            continue

        key = isin or cusip
        meta = ISIN_TO_SLEEVE.get(isin) or CUSIP_TO_META.get(cusip)
        if not meta:
            print(f"⚠ Unknown security: {desc[:50]} (ISIN:{isin} / CUSIP:{cusip})", file=sys.stderr)
            continue

        resolved_isin = meta.get("isin", isin or cusip)
        ter = TER_DATA.get(resolved_isin, {"inst": None, "a": None})

        positions.append({
            "isin": resolved_isin,
            "ticker": meta["ticker"],
            "name": SHORT_NAME.get(resolved_isin, desc),
            "sleeve": meta["sleeve"],
            "value": float(mv),
            "pct": float(pct) if pct else 0,
            "qty": float(qty) if qty else None,
            "price": float(price) if isinstance(price, (int, float)) else None,
            "ter_inst": ter["inst"],
            "ter_a": ter["a"],
            "source": "Pershing",
        })

    return {
        "as_of": as_of,
        "positions": positions,
        "pershing_total": sum(p["value"] for p in positions),
    }


def add_external_alts(pershing_data):
    """Add external alternatives (Carlyle, Flex-Lex, Hamilton Lane)."""
    ext = []
    for alt in EXTERNAL_ALTS:
        if alt.get("value") is None:
            continue  # Skip HL until value confirmed
        ext.append({
            "isin": alt["isin"],
            "ticker": alt["ticker"],
            "name": alt["name"],
            "sleeve": alt["sleeve"],
            "value": alt["value"],
            "pct": None,  # will be recalculated
            "qty": None,
            "price": None,
            "ter_inst": alt.get("ter"),
            "ter_a": None,
            "source": alt["source"],
            "status": alt.get("status"),
        })
    return ext


def recalc_percentages(positions):
    total = sum(p["value"] for p in positions)
    for p in positions:
        p["pct"] = round(p["value"] / total * 100, 2)
    return positions, total


def emit_js_array(positions, as_of=None):
    """Emit BIG_POSITIONS JavaScript literal + POSITIONS_AS_OF marker."""
    sleeve_order = ["Equity", "Alternatives", "Fixed Income", "Cash"]
    positions.sort(key=lambda p: (sleeve_order.index(p["sleeve"]), -p["value"]))
    lines = ["const BIG_POSITIONS = ["]
    for p in positions:
        ter_inst = "null" if p["ter_inst"] is None else f"{p['ter_inst']:.2f}"
        ter_a = "null" if p["ter_a"] is None else f"{p['ter_a']:.2f}"
        lines.append(
            f'    {{ isin: "{p["isin"]}", ticker: "{p["ticker"]}", '
            f'name: "{p["name"]}", sleeve: "{p["sleeve"]}", '
            f'value: {p["value"]:.2f}, pct: {p["pct"]:.2f}, '
            f'terInst: {ter_inst}, terA: {ter_a} }},'
        )
    lines.append("];")
    if as_of:
        # ISO date para el banner de frescura del dashboard
        lines.append("")
        lines.append('// Bumpear POSITIONS_AS_OF al pegar este array en funds_metadata.js')
        lines.append(f'const POSITIONS_AS_OF = "{as_of}";')
    return "\n".join(lines)


def main():
    if len(sys.argv) < 2:
        print("Usage: python pershing_parser.py <path_to_Positions.xlsx>")
        sys.exit(1)

    xlsx = sys.argv[1]
    print(f"[{datetime.now()}] Parsing Pershing: {xlsx}")
    data = parse_pershing(xlsx)
    print(f"\nAs of: {data['as_of']}")
    print(f"Pershing positions: {len(data['positions'])}")
    print(f"Pershing total: ${data['pershing_total']:,.2f}")

    ext = add_external_alts(data)
    print(f"\nExternal alts added: {len(ext)}")
    for e in ext:
        print(f"  + {e['ticker']:8s} ${e['value']:>14,.2f}")

    all_positions = data["positions"] + ext
    all_positions, total_aum = recalc_percentages(all_positions)
    print(f"\n{'='*60}")
    print(f"TOTAL AUM (Pershing + External): ${total_aum:,.2f}")
    print(f"{'='*60}\n")

    # ISO date para POSITIONS_AS_OF (intenta parsear data["as_of"], fallback a hoy)
    try:
        from dateutil import parser as _dtparser
        as_of_iso = _dtparser.parse(data["as_of"]).date().isoformat()
    except Exception:
        as_of_iso = date.today().isoformat()

    js_array = emit_js_array(all_positions, as_of=as_of_iso)
    print(js_array)

    # Save JSON traceability
    out = Path(__file__).parent.parent / "data" / "positions_latest.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump({
            "as_of": data["as_of"],
            "refreshed_at": datetime.now().isoformat(),
            "pershing_total": data["pershing_total"],
            "external_alts_total": sum(e["value"] for e in ext),
            "total_aum": total_aum,
            "positions": all_positions,
        }, f, indent=2)
    print(f"\nJSON written to: {out}")


if __name__ == "__main__":
    main()
