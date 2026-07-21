"""
sync_positions_from_pershing.py
================================
Sincroniza positions_latest.json con los montos actualizados del PDF Pershing
+ CALP manual. Usa Carga Max.xlsx como fuente unificada de montos.

Usage:
    python scripts/sync_positions_from_pershing.py --month 6 --year 2026
"""
import argparse
import json
import re
import shutil
from datetime import datetime, date
from pathlib import Path

import openpyxl

POSITIONS_JSON = Path("C:/Users/lmonp/OneDrive/Desktop/Code/big-dashboard/data/positions_latest.json")
CARGA_MAX = Path("C:/Users/lmonp/Dropbox/BIG/2026/Export % Maximus/Carga Max.xlsx")

# Mapping ISIN entre Listado (Maximus) y positions_latest (dashboard)
# CALP: Listado usa LU2837777825 (código Maximus), dashboard usa LU2827810776 (interno)
ISIN_ALIAS = {
    "LU2837777825": "LU2827810776",  # Carlyle CALP
}

# Ticker (positions_latest) -> ISIN (Listado)
TICKER_TO_ISIN = {
    "CSPX": "IE00B5BMR087",
    "NBGMT": "IE00BFMHRK20",
    "MFSCV": "LU1985812756",
    "THOR": "IE00B6YCBF59",
    "JHGSC": "LU2940405447",
    "LGLI": "IE00BF4KN675",
    "ARGT": "US37950E2596",
    "ILF": "US4642873909",
    "4BRZ": "DE000A0Q4R85",
    "CALP": "LU2837777825",   # el Listado usa el ISIN Maximus
    "IBIT": "US46438F1012",
    "GLD": "US78463V1070",
    "HLEND": "KYG4737U1085",
    "BPCC": "XS2658535526",
    "FLEX": "LU3009548630",  # Franklin Lexington - Listado usa el ISIN Maximus
    "HLGPI": "HLPIF",         # HLPIF es el "ISIN" en Listado
    "GCRED": "GCRED-I",
    "PIMCO-LD": "IE00BDT57R20",
    "PIMCO-INC": "IE00B87KCF77",
    "MANIG": "IE000OE87WX6",
    "PIMCO-EM": "IE00B29K0P99",
    "TGF": "XS2324777171",
    "SGCB": "LU2049315265",
    "MANEM": "IE00089T5MA6",
    "CASH": "CASH",
}


def read_montos_from_carga_max(carga_path: Path) -> dict:
    """Devuelve dict {isin_or_ticker: monto_usd} desde Carga Max.
    Incluye Cash y todos los activos del Listado."""
    wb = openpyxl.load_workbook(str(carga_path), data_only=False)
    ws_in = wb["Input Pershing"]
    ws_list = wb["Listado"]

    result = {}

    # Cash (row 6 del Input Pershing)
    cash = ws_in.cell(row=6, column=8).value
    if isinstance(cash, (int, float)):
        result["CASH"] = cash

    # Activos del Listado (rows 4-27)
    for r in range(4, 28):
        isin = ws_list.cell(row=r, column=1).value
        formula = ws_list.cell(row=r, column=4).value
        if not isin or not formula:
            continue
        m = re.search(r"!I(\d+)\*100", str(formula))
        if not m:
            continue
        src_row = int(m.group(1))
        monto = ws_in.cell(row=src_row, column=8).value
        if isinstance(monto, (int, float)) and monto > 0:
            result[str(isin).strip()] = monto

    return result


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", type=int, required=True)
    ap.add_argument("--year", type=int, required=True)
    ap.add_argument("--carga", type=str, default=str(CARGA_MAX))
    ap.add_argument("--positions", type=str, default=str(POSITIONS_JSON))
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    positions_path = Path(args.positions)
    carga_path = Path(args.carga)

    # Ultimo dia del mes
    from datetime import timedelta
    if args.month == 12:
        last_day = date(args.year, 12, 31)
    else:
        last_day = date(args.year, args.month + 1, 1) - timedelta(days=1)

    print(f"[{datetime.now().isoformat(timespec='seconds')}] Sync positions_latest.json")
    print(f"  As of: {last_day.isoformat()}")
    print(f"  Source: {carga_path.name}")

    # 1) Leer montos del Carga Max
    montos = read_montos_from_carga_max(carga_path)
    print(f"  Montos leidos: {len(montos)}")

    # 2) Leer positions actuales
    positions = json.loads(positions_path.read_text(encoding="utf-8"))
    old_aum = sum(p.get("value", 0) for p in positions["positions"])
    old_asof = positions.get("as_of")
    print(f"  Actual: {old_asof}, AUM ${old_aum:,.2f}, {len(positions['positions'])} positions")

    # 3) Update cada position
    matched = 0
    missing = []
    for p in positions["positions"]:
        isin = p.get("isin", "") or ""
        ticker = p.get("ticker", "") or ""
        # Estrategia:
        # 1) ISIN directo (con alias) en montos
        # 2) Ticker via TICKER_TO_ISIN -> ISIN listado -> montos
        # 3) Ticker directo en montos
        key = ISIN_ALIAS.get(isin, isin) if isin else None
        monto = montos.get(key) if key else None
        if monto is None and ticker in TICKER_TO_ISIN:
            monto = montos.get(TICKER_TO_ISIN[ticker])
        if monto is None and ticker:
            monto = montos.get(ticker)
        if monto is not None:
            p["value"] = round(monto, 2)
            p["price_as_of"] = last_day.isoformat()
            # Recalcular price si tiene qty
            if isinstance(p.get("qty"), (int, float)) and p["qty"] > 0:
                p["price"] = round(monto / p["qty"], 4)
            matched += 1
        else:
            missing.append((isin, ticker, p.get("name", "")))

    print(f"  Matched: {matched}/{len(positions['positions'])}")
    if missing:
        print(f"  MISSING (mantienen su value viejo):")
        for isin, ticker, name in missing:
            print(f"    - {isin} {ticker} {name}")

    # 4) Update metadata
    new_aum = sum(p["value"] for p in positions["positions"])
    positions["as_of"] = f"{last_day.strftime('%b %d, %Y')}"
    positions["as_of_date"] = last_day.isoformat()
    positions["refreshedAt"] = datetime.now().isoformat()
    positions["total_aum"] = round(new_aum, 2)
    positions["n_positions"] = len(positions["positions"])

    print()
    print(f"  === Resumen ===")
    print(f"  as_of:       {old_asof} -> {positions['as_of']}")
    print(f"  total_aum:   ${old_aum:,.2f} -> ${new_aum:,.2f}")

    if args.dry_run:
        print(f"  DRY RUN - no se guarda")
        return

    # Backup + save
    backup = positions_path.parent / f"{positions_path.stem}.{datetime.now().strftime('%Y%m%d-%H%M%S')}.bak.json"
    shutil.copy2(positions_path, backup)
    print(f"  Backup: {backup.name}")

    positions_path.write_text(
        json.dumps(positions, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"  Guardado: {positions_path.name}")


if __name__ == "__main__":
    main()
