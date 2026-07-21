"""
Genera hoja para Maximus con Fila 11 (Date) + Fila 36 (Gross NAV) desde el
cierre del mes anterior al cierre del mes actual.

Uso mensual: al cierre de cada mes, correr este script cambiando MONTH_YEAR.
Output: nuevo archivo Excel en Desktop con la hoja lista para copiar/pegar.
"""
import json
from datetime import date, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from copy import copy

# ============ CONFIG (cambiar mensualmente) ============
YEAR = 2026
MONTH = 6  # mes de cierre
# =======================================================

# Rango: último día del mes anterior → último día del mes actual
if MONTH == 1:
    START = date(YEAR - 1, 12, 31)
else:
    # Ultimo día del mes anterior
    prev = date(YEAR, MONTH, 1) - timedelta(days=1)
    START = prev

# Último día del mes actual
if MONTH == 12:
    END = date(YEAR, 12, 31)
else:
    END = date(YEAR, MONTH + 1, 1) - timedelta(days=1)

print(f"Rango: {START} -> {END}")

# ============ LEER LYNK NAV SERIES ============
lynk_file = Path("C:/Users/lmonp/OneDrive/Desktop/Code/big-dashboard/data/lynk_nav_series.json")
d = json.load(open(lynk_file, encoding="utf-8"))
series = d.get("navSeries") or d.get("series") or []
nav_by_date = {}
for p in series:
    dt = p.get("date")
    nav = p.get("value") or p.get("nav")
    if dt and nav is not None:
        nav_by_date[dt] = float(nav)

# Construir mapa día a día con carry-forward para weekends/feriados.
# Pre-cargar last_nav con el NAV más reciente ANTES de START (útil si START cae en weekend)
last_nav = None
for check_date in sorted(nav_by_date.keys(), reverse=True):
    if check_date <= START.isoformat():
        last_nav = nav_by_date[check_date]
        print(f"Anchor pre-START: {check_date} NAV {last_nav}")
        break

all_dates = []
navs = []
curr = START
while curr <= END:
    iso = curr.isoformat()
    all_dates.append(curr)
    if iso in nav_by_date:
        last_nav = nav_by_date[iso]
    navs.append(last_nav)
    curr += timedelta(days=1)

print(f"Total dias: {len(all_dates)} | NAVs mapeados: {sum(1 for n in navs if n is not None)}")

# ============ CARGAR EL EXCEL ORIGINAL Y AGREGAR HOJA ============
src = Path("C:/Users/lmonp/Downloads/ProCapital_XS3037627794_LS104 (5).xlsx")
out = Path(f"C:/Users/lmonp/OneDrive/Desktop/ProCapital_LS104_NAV_{YEAR}-{MONTH:02d}.xlsx")

wb = openpyxl.load_workbook(str(src))
sheet_name = f"NAV {START.strftime('%b-%Y')} to {END.strftime('%b-%Y')}"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
ws = wb.create_sheet(sheet_name, index=0)

# Copiar header rows 1-7 (metadata Series Name, ISIN, etc)
src_ws = wb["NAV Calculation"]
for r in range(1, 8):
    for c in range(1, 3):
        v = src_ws.cell(row=r, column=c).value
        ws.cell(row=r, column=c).value = v

# Estilos
bold = Font(bold=True)
date_fill = PatternFill("solid", fgColor="FFFF00")  # amarillo (fila 11)
nav_fill = PatternFill("solid", fgColor="C00000")   # rojo (fila 36)
white_bold = Font(bold=True, color="FFFFFF")

# Fila 11: Date
ws.cell(row=11, column=1).value = "Date"
ws.cell(row=11, column=1).font = bold
ws.cell(row=11, column=1).fill = date_fill

# Fila 57: NAV
ws.cell(row=57, column=1).value = "NAV"
ws.cell(row=57, column=1).font = white_bold
ws.cell(row=57, column=1).fill = nav_fill

# Llenar columnas C en adelante
for i, (dt, nav) in enumerate(zip(all_dates, navs)):
    col = 3 + i  # empieza en col C
    # Row 11: fecha
    date_cell = ws.cell(row=11, column=col)
    date_cell.value = dt
    date_cell.number_format = "m/d/yyyy"
    date_cell.font = bold
    date_cell.fill = date_fill
    date_cell.alignment = Alignment(horizontal="center")
    # Row 36: NAV
    nav_cell = ws.cell(row=57, column=col)
    if nav is not None:
        nav_cell.value = nav
        nav_cell.number_format = "0.000"
    nav_cell.font = white_bold
    nav_cell.fill = nav_fill
    nav_cell.alignment = Alignment(horizontal="center")

# ============ COLUMNAS EXTRAS: Return Mes + YTD ============
nav_ini = navs[0] if navs else None
nav_fin = navs[-1] if navs else None
return_mes = ((nav_fin / nav_ini) - 1) * 100 if (nav_ini and nav_fin) else None

# YTD anchor: NAV 12/31 del año anterior
ytd_anchor_date = date(YEAR - 1, 12, 31).isoformat()
nav_ytd_anchor = nav_by_date.get(ytd_anchor_date)
if nav_ytd_anchor is None:
    for check in sorted(nav_by_date.keys(), reverse=True):
        if check <= ytd_anchor_date:
            nav_ytd_anchor = nav_by_date[check]
            break
return_ytd = ((nav_fin / nav_ytd_anchor) - 1) * 100 if (nav_ytd_anchor and nav_fin) else None

print(f"\nReturn Mes ({START} -> {END}): {return_mes:.2f}%" if return_mes is not None else "Return Mes: N/A")
print(f"Return YTD ({ytd_anchor_date} -> {END}): {return_ytd:.2f}%" if return_ytd is not None else "Return YTD: N/A")

# Colocar en 2 cols nuevas al final
last_col = 3 + len(all_dates) - 1
gold_fill = PatternFill("solid", fgColor="D4AF37")
green_fill = PatternFill("solid", fgColor="70AD47")

col_mes = last_col + 2
ws.cell(row=11, column=col_mes).value = "Return Mes"
ws.cell(row=11, column=col_mes).font = Font(bold=True, color="FFFFFF")
ws.cell(row=11, column=col_mes).fill = gold_fill
ws.cell(row=11, column=col_mes).alignment = Alignment(horizontal="center")
mes_cell = ws.cell(row=57, column=col_mes)
if return_mes is not None:
    mes_cell.value = return_mes / 100
    mes_cell.number_format = "0.00%"
mes_cell.font = Font(bold=True, color="FFFFFF")
mes_cell.fill = gold_fill
mes_cell.alignment = Alignment(horizontal="center")

col_ytd = col_mes + 1
ws.cell(row=11, column=col_ytd).value = "Return YTD"
ws.cell(row=11, column=col_ytd).font = Font(bold=True, color="FFFFFF")
ws.cell(row=11, column=col_ytd).fill = green_fill
ws.cell(row=11, column=col_ytd).alignment = Alignment(horizontal="center")
ytd_cell = ws.cell(row=57, column=col_ytd)
if return_ytd is not None:
    ytd_cell.value = return_ytd / 100
    ytd_cell.number_format = "0.00%"
ytd_cell.font = Font(bold=True, color="FFFFFF")
ytd_cell.fill = green_fill
ytd_cell.alignment = Alignment(horizontal="center")

# Ajustar anchos
ws.column_dimensions["A"].width = 22
for i in range(len(all_dates)):
    col_letter = openpyxl.utils.get_column_letter(3 + i)
    ws.column_dimensions[col_letter].width = 12
ws.column_dimensions[openpyxl.utils.get_column_letter(col_mes)].width = 14
ws.column_dimensions[openpyxl.utils.get_column_letter(col_ytd)].width = 14

wb.save(str(out))
print(f"\nOK Excel guardado: {out}")
print(f"   Hoja creada: '{sheet_name}'")
print(f"   Fila 11 (Date): {len(all_dates)} fechas de {START} a {END}")
print(f"   Fila 36 (Gross NAV): {sum(1 for n in navs if n is not None)} NAVs (con carry-forward para weekends)")
print()
print("Preview:")
for i in range(min(5, len(all_dates))):
    print(f"   {all_dates[i]}: NAV {navs[i]}")
print("   ...")
for i in range(max(0, len(all_dates)-3), len(all_dates)):
    print(f"   {all_dates[i]}: NAV {navs[i]}")
