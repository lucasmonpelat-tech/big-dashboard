"""
refresh_analyzer.py
===================
Actualiza pampa-big-analyzer.html con:
- Header con fecha del mes actual del cierre
- Array FUND[] con pesos al cierre desde Carga Max.xlsx

Usage:
    python scripts/refresh_analyzer.py --month 6 --year 2026
"""
import argparse
import re
from datetime import datetime, date
from pathlib import Path
import openpyxl

CARGA_MAX = Path("C:/Users/lmonp/Dropbox/BIG/2026/Export % Maximus/Carga Max.xlsx")
ANALYZER_HTML = Path("C:/Users/lmonp/OneDrive/Desktop/Code/big-dashboard/pampa-big-analyzer.html")

MONTH_ABBR_ES = {
    1: "ene", 2: "feb", 3: "mar", 4: "abr", 5: "may", 6: "jun",
    7: "jul", 8: "ago", 9: "sep", 10: "oct", 11: "nov", 12: "dic",
}


def read_pesos_from_carga_max(carga_path: Path) -> list:
    """Lee Carga Max.xlsx hoja Listado + Input Pershing y devuelve list de
    {pct, isin, name, clase} para el array FUND del analyzer.
    Excluye Cash (row 6 del Input Pershing)."""
    # data_only=False para poder leer las formulas del Listado col D.
    # Los montos raw del Input Pershing col H se guardan como valores, no formulas.
    wb = openpyxl.load_workbook(str(carga_path), data_only=False)
    ws_listado = wb["Listado"]
    ws_input = wb["Input Pershing"]

    # Construir dict {ref_row: (monto, name)} desde Input Pershing (rows 6-30)
    input_data = {}
    total = 0
    for r in range(6, 31):
        monto = ws_input.cell(row=r, column=8).value
        name = ws_input.cell(row=r, column=5).value
        if isinstance(monto, (int, float)) and monto > 0:
            input_data[r] = (monto, name)
            total += monto
    if not total:
        raise ValueError("No hay montos en Input Pershing H6:H30")

    # Mapping clase de activo desde la col D (Tipo de Activo)
    def clase_from_tipo(tipo: str) -> str:
        if not tipo:
            return "Renta Variable"
        tipo_l = tipo.lower()
        if "cash" in tipo_l or "money" in tipo_l:
            return "Cash"
        if "fixed" in tipo_l or "debt" in tipo_l:
            return "Renta Fija"
        if "mutual" in tipo_l:
            return None  # se decide por el propio activo
        if "exchange" in tipo_l or "etf" in tipo_l:
            return None
        if "alternative" in tipo_l:
            return "Alternativos"
        return None

    # Ir por el Listado — cada row tiene una fórmula que apunta al Input Pershing.
    # Como openpyxl no evalua formulas, la parseamos manualmente para saber que
    # row del Input Pershing referencia, y calculamos pct = monto/total*100.
    funds = []
    for r in range(4, 28):  # rows 4-27 del Listado
        isin = ws_listado.cell(row=r, column=1).value
        name = ws_listado.cell(row=r, column=2).value
        formula = ws_listado.cell(row=r, column=4).value
        if not isin or not name or not formula:
            continue
        # Formula: ='Input Pershing'!I<row>*100 -> extraemos <row>
        m = re.search(r"!I(\d+)\*100", str(formula))
        if not m:
            # Fallback: puede ser valor numerico directo
            if isinstance(formula, (int, float)):
                pct = formula
            else:
                continue
        else:
            src_row = int(m.group(1))
            if src_row not in input_data:
                continue
            monto, _ = input_data[src_row]
            pct = monto / total * 100
        if pct <= 0:
            continue
        # Determinar clase por nombre
        name_l = str(name).lower()
        clase = "Renta Variable"
        if any(x in name_l for x in ["pimco", "man glg", "tenac", "schroder", "man em"]):
            clase = "Renta Fija"
        elif any(x in name_l for x in ["carlyle", "bitcoin", "gold", "hps", "barings", "franklin lexington",
                                        "hamilton lane", "golub"]):
            clase = "Alternativos"
        funds.append({
            "pct": round(pct, 2),
            "isin": str(isin).strip(),
            "name": str(name).strip(),
            "clase": clase,
        })

    return funds, total


def format_fund_array(funds: list) -> str:
    """Genera el array FUND[] como texto JS ordenado por clase (RV, RF, Alts)."""
    order = ["Renta Variable", "Renta Fija", "Alternativos"]
    lines = []
    for cls in order:
        for f in funds:
            if f["clase"] != cls:
                continue
            lines.append(
                f"    {{pct:{f['pct']:>5}, isin:'{f['isin']}', name:'{f['name'].replace(chr(39), chr(92)+chr(39))}', clase:'{f['clase']}'}},"
            )
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", type=int, required=True, help="Mes (1-12)")
    ap.add_argument("--year", type=int, required=True)
    ap.add_argument("--carga", type=str, default=str(CARGA_MAX))
    ap.add_argument("--html", type=str, default=str(ANALYZER_HTML))
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    carga_path = Path(args.carga)
    html_path = Path(args.html)

    # Ultimo dia del mes
    if args.month == 12:
        last_day = date(args.year, 12, 31)
    else:
        from datetime import timedelta
        last_day = date(args.year, args.month + 1, 1) - timedelta(days=1)

    fecha_label = f"Al {last_day.day} {MONTH_ABBR_ES[args.month]} {args.year}"
    print(f"[{datetime.now().isoformat(timespec='seconds')}] Refresh analyzer HTML")
    print(f"  Cierre: {last_day.isoformat()} ({fecha_label})")
    print(f"  Carga Max: {carga_path.name}")
    print(f"  Analyzer:  {html_path.name}")

    # 1) Leer pesos
    funds, total = read_pesos_from_carga_max(carga_path)
    print(f"  Activos leidos: {len(funds)}")
    print(f"  Total portfolio: ${total:,.2f}")
    print(f"  Distribucion por clase:")
    from collections import Counter
    counts = Counter(f["clase"] for f in funds)
    for cls, n in counts.items():
        weight = sum(f["pct"] for f in funds if f["clase"] == cls)
        print(f"    {cls}: {n} activos ({weight:.1f}%)")

    # 2) Leer HTML actual
    html = html_path.read_text(encoding="utf-8")

    # 3) Reemplazar header "Al DD mmm YYYY"
    new_header_line = f"Currency Exposure &amp; Current Yield Analyzer &middot; {len(funds)} activos &middot; {fecha_label}"
    html_new, n1 = re.subn(
        r"Currency Exposure &amp; Current Yield Analyzer.*?Al \d+ [a-z]+ \d{4}",
        new_header_line,
        html,
    )
    if n1 == 0:
        # Fallback: alternate patterns
        html_new, n1 = re.subn(
            r"(?:Currency Exposure).*?(?:activos).*?Al \d+ [a-z]+ \d{4}",
            new_header_line,
            html,
        )
    print(f"  Header updates: {n1}")

    # 4) Reemplazar array FUND[]
    new_fund_block = f"const FUND = [\n{format_fund_array(funds)}\n]"
    html_new, n2 = re.subn(
        r"const FUND = \[.*?\]",
        new_fund_block,
        html_new,
        count=1,
        flags=re.DOTALL,
    )
    print(f"  FUND[] updates: {n2}")

    if n1 == 0 or n2 == 0:
        print(f"  WARN: no se reemplazo todo. Revisar patterns.")

    if args.dry_run:
        print(f"  DRY RUN - no se guarda")
        # Preview del FUND[]
        print()
        print("PREVIEW FUND[]:")
        print(new_fund_block[:1500])
        return

    # 5) Backup + save
    backup = html_path.parent / f"{html_path.stem}.{datetime.now().strftime('%Y%m%d-%H%M%S')}.bak"
    backup.write_text(html, encoding="utf-8")
    print(f"  Backup: {backup.name}")

    html_path.write_text(html_new, encoding="utf-8")
    print(f"  Guardado: {html_path.name}")

    print()
    print(f"  Siguiente paso: abrir {html_path.name} en el browser")
    print(f"    1. Click en '▶ Ejecutar análisis de currency exposure' (Tab 1)")
    print(f"    2. Click en '▶ Ejecutar búsqueda de Current Yield' (Tab 2)")
    print(f"    3. Tab 3 (Resumen consolidado) muestra los 2 numeros para el Factsheet")


if __name__ == "__main__":
    main()
