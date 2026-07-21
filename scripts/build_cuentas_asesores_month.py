"""
build_cuentas_asesores_month.py
================================
Paso 2 del cierre mensual BIG.

Agrega hoja nueva del mes en BIG Cuentas - Asesores.xlsx copiando la ultima
hoja mensual (ej MAY) como template. Reemplaza data con positions Pershing
del mes nuevo, actualiza header con PAMPA total, y agrega col Lucas (SUMIF).
El Mapping se actualiza con Azul Fernandez -> Lucas si aun no esta.

Usage:
    python scripts/build_cuentas_asesores_month.py --month JUN --pampa 28212.01
"""
import argparse
import shutil
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

FOLDER = Path("C:/Users/lmonp/Dropbox/BIG/2026/Luiquidacion Comisiones BIG")
DEFAULT_CUENTAS = FOLDER / "BIG Cuentas - Asesores (1).xlsx"

MONTH_TO_ABBR = {
    1: "ENE", 2: "FEB", 3: "MAR", 4: "ABR", 5: "MAY",
    6: "JUN", 7: "JUL", 8: "AGO", 9: "SEP",
    10: "OCT", 11: "NOV", 12: "DIC",
}
MONTH_ABBR_TO_INT = {v: k for k, v in MONTH_TO_ABBR.items()}

# Asesores activos (Alan y Gonzalo ya no estan). Orden importa: I=Fernando, J=Segundo, K=Fede, L=Lucas.
ASESORES = ["Fernando", "Segundo", "Fede", "Lucas"]

# Nueva cuenta a mapear a Lucas (confirmada por Lucas 2026-07-03)
LUCAS_ACCOUNTS = {
    "JXD113179": "Lucas",  # Azul Fernandez
}


def find_positions_file(folder: Path, month_int: int, year: int) -> Path:
    month_names_full = {
        1: ["Enero", "Ene"], 2: ["Febrero", "Feb"], 3: ["Marzo", "Mar"],
        4: ["Abril", "Abr"], 5: ["Mayo", "May"], 6: ["Junio", "Jun"],
        7: ["Julio", "Jul"], 8: ["Agosto", "Ago"], 9: ["Septiembre", "Sep"],
        10: ["Octubre", "Oct"], 11: ["Noviembre", "Nov"], 12: ["Diciembre", "Dic"],
    }
    candidates = [p for p in folder.glob("Positions_JXD*.xlsx") if not p.name.startswith("~$")]
    for name in month_names_full[month_int]:
        for path in candidates:
            if name.lower() in path.name.lower():
                return path
    if len(candidates) == 1:
        return candidates[0]
    raise FileNotFoundError(
        f"No encuentro Positions_JXD para el mes {month_int}. "
        f"Candidatos: {[p.name for p in candidates]}"
    )


def read_positions(path: Path):
    """Lee export Pershing y devuelve lista de {account, name, mv}."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    header_row = None
    for r in range(1, 50):
        if ws.cell(row=r, column=1).value == "Account #":
            header_row = r
            break
    if header_row is None:
        raise ValueError("No encuentro header 'Account #' en Positions_JXD")

    col_acc = col_name = col_mv = None
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=header_row, column=c).value
        if h == "Account #":
            col_acc = c
        elif h == "Short Name":
            col_name = c
        elif h and "Market Value" in str(h):
            col_mv = c
    if not all([col_acc, col_name, col_mv]):
        raise ValueError(f"Cols no detectadas: acc={col_acc} name={col_name} mv={col_mv}")

    positions = []
    for r in range(header_row + 1, ws.max_row + 1):
        acc = ws.cell(row=r, column=col_acc).value
        name = ws.cell(row=r, column=col_name).value
        mv = ws.cell(row=r, column=col_mv).value
        if not acc or not str(acc).startswith("JXD"):
            continue
        if not isinstance(mv, (int, float)):
            continue
        positions.append({
            "account": str(acc).strip(),
            "name": str(name or "").strip(),
            "mv": float(mv),
        })
    return positions


def ensure_mapping(wb, extra: dict):
    """Agrega entries al Mapping sheet si no existen. Devuelve list de agregados."""
    ws = wb["Mapping"]
    existing = {}
    max_r = ws.max_row
    for r in range(2, max_r + 1):
        acc = ws.cell(row=r, column=1).value
        ase = ws.cell(row=r, column=2).value
        if acc:
            existing[str(acc).strip()] = ase
    added = []
    next_row = max_r + 1
    for acc, asesor in extra.items():
        if acc in existing:
            if existing[acc] != asesor:
                print(f"    WARN: {acc} ya esta en Mapping como '{existing[acc]}', no como '{asesor}'")
            continue
        ws.cell(row=next_row, column=1).value = acc
        ws.cell(row=next_row, column=2).value = asesor
        added.append((acc, asesor))
        next_row += 1
    return added


def backup_file(path: Path):
    backup_dir = path.parent / ".backups"
    backup_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = backup_dir / f"{path.stem}.{ts}{path.suffix}"
    shutil.copy2(path, backup_path)
    return backup_path


def get_template_sheet(wb):
    """Devuelve nombre de la ultima hoja mensual como template."""
    for name in reversed(wb.sheetnames):
        if name in MONTH_ABBR_TO_INT:
            return name
    for name in reversed(wb.sheetnames):
        if name.startswith("Com "):
            return name
    raise ValueError("No hay hoja mensual template")


def parse_assign(assign_str):
    """Parsea '--assign JXD1=Fernando,JXD2=Lucas' -> {'JXD1': 'Fernando', 'JXD2': 'Lucas'}"""
    if not assign_str:
        return {}
    result = {}
    for pair in assign_str.split(","):
        pair = pair.strip()
        if not pair:
            continue
        if "=" not in pair:
            raise ValueError(f"Assign invalido (falta '='): {pair}")
        acc, asesor = pair.split("=", 1)
        result[acc.strip()] = asesor.strip()
    return result


def scan_orphans(positions, mapping, extra_assign):
    """Devuelve list de huerfanos: (account, name, mv) presentes en positions
    pero NO en mapping (y no cubiertos por LUCAS_ACCOUNTS ni extra_assign)."""
    covered = set(mapping.keys()) | set(LUCAS_ACCOUNTS.keys()) | set(extra_assign.keys())
    orphans = []
    for p in positions:
        if p["account"] not in covered:
            orphans.append((p["account"], p["name"], p["mv"]))
    return orphans


def load_mapping_dict(cuentas_path):
    """Devuelve dict {account: asesor} desde la hoja Mapping."""
    wb = openpyxl.load_workbook(str(cuentas_path), data_only=True)
    ws = wb["Mapping"]
    mapping = {}
    for r in range(2, ws.max_row + 1):
        acc = ws.cell(row=r, column=1).value
        ase = ws.cell(row=r, column=2).value
        if acc:
            mapping[str(acc).strip()] = ase
    return mapping


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", type=str, required=True, help="Abbr JUN/JUL/etc")
    ap.add_argument("--year", type=int, default=datetime.now().year)
    ap.add_argument("--pampa", type=float, default=None, help="PAMPA total del mes (obligatorio salvo --check-only)")
    ap.add_argument("--positions", type=str, default=None)
    ap.add_argument("--cuentas", type=str, default=str(DEFAULT_CUENTAS))
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--force", action="store_true")
    ap.add_argument("--check-only", action="store_true", help="Solo pre-scan de huerfanos, no modifica nada")
    ap.add_argument("--assign", type=str, default="", help="Asignar cuentas huerfanas: 'JXD1=Fernando,JXD2=Lucas'")
    ap.add_argument("--allow-orphans", action="store_true", help="Continuar aunque haya huerfanos (aparecen FALTA EN MAPPING)")
    args = ap.parse_args()

    month_abbr = args.month.upper()
    if month_abbr not in MONTH_ABBR_TO_INT:
        raise ValueError(f"Mes invalido: {args.month}")
    month_int = MONTH_ABBR_TO_INT[month_abbr]

    cuentas_path = Path(args.cuentas)
    if not cuentas_path.exists():
        raise FileNotFoundError(f"No encuentro: {cuentas_path}")

    if args.positions:
        positions_path = Path(args.positions)
    else:
        positions_path = find_positions_file(cuentas_path.parent, month_int, args.year)
    if not positions_path.exists():
        raise FileNotFoundError(f"No encuentro: {positions_path}")

    print(f"[{datetime.now().isoformat(timespec='seconds')}] Paso 2: Cuentas Asesores {month_abbr} {args.year}")
    print(f"  Positions: {positions_path.name}")

    positions = read_positions(positions_path)
    total_mv = sum(p["mv"] for p in positions)
    print(f"  Comitentes: {len(positions)} | Total MV: ${total_mv:,.2f}")

    # === PRE-SCAN: detectar huerfanos ANTES de modificar nada ===
    extra_assign = parse_assign(args.assign)
    mapping_current = load_mapping_dict(cuentas_path)
    orphans = scan_orphans(positions, mapping_current, extra_assign)

    print()
    print(f"  === Pre-scan Mapping ===")
    print(f"  Cuentas en Mapping actual: {len(mapping_current)}")
    print(f"  Cuentas hardcoded (Lucas): {len(LUCAS_ACCOUNTS)}")
    print(f"  Cuentas asignadas via --assign: {len(extra_assign)}")
    print(f"  HUERFANOS (sin asignar): {len(orphans)}")
    if orphans:
        pampa_est = args.pampa or 0
        for acc, name, mv in orphans:
            share = mv / total_mv * 100
            trailer_est = mv / total_mv * pampa_est
            print(f"    - {acc} {name:15s} MV=${mv:>12,.2f}  ({share:.1f}%)  trailer_est=${trailer_est:,.2f}")

    if args.check_only:
        print()
        print(f"  --check-only activado: NO se modifica nada.")
        if orphans:
            print(f"  Para resolver: correr con --assign 'JXD1=Fernando,JXD2=Lucas,...'")
        return orphans

    if args.pampa is None:
        print(f"  ERROR: --pampa es obligatorio (viene del Paso 1)")
        return

    if orphans and not args.allow_orphans:
        print()
        print(f"  ERROR: {len(orphans)} cuenta(s) huerfana(s) sin asignar.")
        print(f"  Opciones:")
        print(f"    A) Correr con --assign 'JXD1=Fernando,JXD2=Lucas,...' para agregarlas al Mapping")
        print(f"    B) Correr con --allow-orphans para dejarlas como FALTA EN MAPPING (no recomendado)")
        return

    print(f"  PAMPA total del mes: ${args.pampa:,.2f}")

    wb = openpyxl.load_workbook(str(cuentas_path))

    if month_abbr in wb.sheetnames:
        if args.force:
            del wb[month_abbr]
        else:
            print(f"  ERROR: Hoja '{month_abbr}' ya existe. Usar --force para sobreescribir.")
            return

    template_name = get_template_sheet(wb)
    print(f"  Template: '{template_name}'")

    if not args.dry_run:
        bpath = backup_file(cuentas_path)
        print(f"  Backup: {bpath.name}")

    # Agregar al Mapping: LUCAS_ACCOUNTS (hardcoded) + --assign (dinamico)
    to_add = {**LUCAS_ACCOUNTS, **extra_assign}
    added = ensure_mapping(wb, to_add)
    if added:
        for acc, asesor in added:
            print(f"  Mapping: agregado {acc} -> {asesor}")

    src = wb[template_name]
    ws_new = wb.copy_worksheet(src)
    ws_new.title = month_abbr

    # Row 1 header: H1="Trailer {MES}", I1=PAMPA, J1="Check", K1=formula check
    ws_new.cell(row=1, column=8).value = f"Trailer {month_abbr}"
    ws_new.cell(row=1, column=9).value = args.pampa
    ws_new.cell(row=1, column=9).number_format = "$#,##0.00"
    ws_new.cell(row=1, column=10).value = "Check"
    # Formula check extendida a L (nueva col Lucas)
    ws_new.cell(row=1, column=11).value = '=IF(ROUND(SUM(I3:L3)-I1,2)=0,"OK","REVISAR")'

    # Row 2 headers de asesores (I=Fernando, J=Segundo, K=Fede, L=Lucas)
    for i, asesor in enumerate(ASESORES):
        col_idx = 9 + i
        ws_new.cell(row=2, column=col_idx).value = asesor

    # Limpiar rows 3+ existentes
    max_row_current = ws_new.max_row
    for r in range(3, max_row_current + 1):
        for c in range(1, 13):
            ws_new.cell(row=r, column=c).value = None

    # Pegar positions
    n = len(positions)
    total_row = 3 + n

    for i, p in enumerate(positions):
        r = 3 + i
        ws_new.cell(row=r, column=1).value = p["account"]
        ws_new.cell(row=r, column=2).value = p["name"]
        ws_new.cell(row=r, column=3).value = p["mv"]
        ws_new.cell(row=r, column=3).number_format = "$#,##0.00"
        ws_new.cell(row=r, column=4).value = f"=C{r}/$C${total_row}"
        ws_new.cell(row=r, column=4).number_format = "0.00%"
        ws_new.cell(row=r, column=5).value = f'=IFERROR(VLOOKUP(A{r},Mapping!$A:$B,2,FALSE),"FALTA EN MAPPING")'
        ws_new.cell(row=r, column=6).value = f"=+D{r}*$I$1"
        ws_new.cell(row=r, column=6).number_format = "$#,##0.00"

    # SUMIF asesores solo en row 3
    for i, asesor in enumerate(ASESORES):
        col_idx = 9 + i
        ws_new.cell(row=3, column=col_idx).value = f'=SUMIF(E:E,"{asesor}",F:F)'
        ws_new.cell(row=3, column=col_idx).number_format = "$#,##0.00"

    # Row total
    ws_new.cell(row=total_row, column=1).value = "TOTAL"
    ws_new.cell(row=total_row, column=3).value = f"=SUM(C3:C{total_row-1})"
    ws_new.cell(row=total_row, column=3).number_format = "$#,##0.00"
    ws_new.cell(row=total_row, column=4).value = f"=SUM(D3:D{total_row-1})"
    ws_new.cell(row=total_row, column=4).number_format = "0.00%"
    ws_new.cell(row=total_row, column=6).value = f"=SUM(F3:F{total_row-1})"
    ws_new.cell(row=total_row, column=6).number_format = "$#,##0.00"

    if args.dry_run:
        print("  DRY RUN - no se guarda")
        return
    wb.save(str(cuentas_path))
    print(f"  Guardado: {cuentas_path.name}")

    # Reporte: split calculado manualmente
    print()
    print(f"  === Split {month_abbr} {args.year} ===")
    wb_calc = openpyxl.load_workbook(str(cuentas_path), data_only=True)
    ws_map = wb_calc["Mapping"]
    mapping = {}
    for r in range(2, ws_map.max_row + 1):
        acc = ws_map.cell(row=r, column=1).value
        ase = ws_map.cell(row=r, column=2).value
        if acc:
            mapping[str(acc).strip()] = ase

    split = {a: 0.0 for a in ASESORES}
    faltantes = []
    for p in positions:
        share = p["mv"] / total_mv
        trailer = share * args.pampa
        ase = mapping.get(p["account"])
        if ase in split:
            split[ase] += trailer
        else:
            faltantes.append((p["account"], p["name"], ase, trailer))

    total_check = 0
    for asesor in ASESORES:
        v = split[asesor]
        total_check += v
        pct = v / args.pampa * 100 if args.pampa else 0
        print(f"    {asesor:10s}: ${v:>10,.2f}  ({pct:>5.1f}%)")
    otros = args.pampa - total_check
    print(f"    {'Otros/NA':10s}: ${otros:>10,.2f}  ({otros/args.pampa*100:>5.1f}%)")
    print(f"    {'TOTAL':10s}: ${args.pampa:>10,.2f}")
    print(f"    Faltantes (Alan/no mapeado): {len(faltantes)} cuentas")
    if faltantes and len(faltantes) <= 10:
        for acc, name, ase, tr in faltantes[:10]:
            print(f"       {acc} {name!s:15s} -> {ase!r} (${tr:,.2f})")


if __name__ == "__main__":
    main()
