"""
build_carga_maximus.py
======================
Skill /carga-maximus - actualiza el Excel "Carga Max.xlsx" con posiciones
del PDF Pershing statement mensual, listo para subir a Maximus.

Flujo:
1. Parsea PDF Pershing (extrae 23 posiciones con Security ID + Market Value + Cash)
2. Valida total contra "Ending Account Value" del statement
3. Mapea cada SecID al row correspondiente en Input Pershing (rows 6-29)
4. Escribe montos en col H
5. Agrega CALP manual en row 30 (Carlyle AlpInvest, Private Equity, no viene en Pershing)
6. Las fórmulas de % (col I) y de Listado (col D) se recalculan solas
7. Guarda in-place con backup

Usage:
    python scripts/build_carga_maximus.py --pdf "path/al/statement.pdf"
    python scripts/build_carga_maximus.py --pdf "..." --calp 2738984
"""
import argparse
import re
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
import pdfplumber

FOLDER = Path("C:/Users/lmonp/Dropbox/BIG/2026/Export % Maximus")
DEFAULT_EXCEL = FOLDER / "Carga Max.xlsx"

# Valor default del CALP (Carlyle AlpInvest). Actualizar cuando llega el statement Carlyle trimestral.
DEFAULT_CALP = 2_738_984.00

# Mapping SecID (Pershing) -> row en hoja "Input Pershing" del Excel
# Las rows son fijas — la hoja Listado usa fórmulas hardcoded (ej: R11 CSPX <- I23)
SEC_ID_TO_ROW = {
    # Fixed Income
    "G7151GAA7": 7,   # Barings BPCC
    "G5478EAA2": 8,   # LFEEDER / Tenac
    # Mutual Funds
    "G5S05E528": 9,   # Lazard Global Listed Infra
    "G5896H580": 10,  # Man GLG IG
    "G594CD616": 11,  # Man EM Corp Credit
    "G6430T809": 12,  # NB Megatrends
    "G7S11T150": 13,  # PIMCO Low Duration
    "G7097Y503": 14,  # PIMCO EM Local Bond
    "G7113P361": 15,  # PIMCO Income
    "G8T49N214": 16,  # Thornburg Income Builder
    "L4R58Q111": 17,  # Franklin Lexington PE Secondaries
    "L468AA656": 18,  # Janus Henderson Global Smaller
    "L4680C117": 19,  # Hamilton Lane Global Private Infra
    "L6366L196": 20,  # MFS Contrarian Value
    "L8147L735": 21,  # Schroder GAIA Cat Bond
    # ETFs
    "IDMBF": 22,      # iShares MSCI Brazil UCITS
    "CSTNL": 23,      # CSPX (iShares Core S&P 500 UCITS)
    "ARGT": 24,       # Global X MSCI Argentina
    "ILF": 25,        # iShares LatAm 40
    "IBIT": 26,       # iShares Bitcoin Trust
    "GLD": 27,        # SPDR Gold Shares
    # Alternative Investments
    "379LP0083": 28,  # Golub Capital GCRED
    "449LP0055": 29,  # HPS Corporate Lending
}
CASH_ROW = 6
CALP_ROW = 30
TOTAL_ROW = 31


def parse_pershing_pdf(pdf_path: Path) -> dict:
    """Devuelve dict con positions (list), cash, ending_value.

    positions: [{'sec_id', 'name', 'market_value'}]
    """
    with pdfplumber.open(str(pdf_path)) as pdf:
        full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    positions = []
    seen = set()
    lines = full_text.split("\n")
    for i, line in enumerate(lines):
        m = re.match(r"\s*Security Identifier:\s*(\S+)", line)
        if not m:
            continue
        sec_id = m.group(1)
        if sec_id in seen:
            continue
        # Buscar hacia atras (max 5 lineas) la linea con quantity + price + MV (o quantity + N/A + MV)
        for j in range(i - 1, max(-1, i - 6), -1):
            prev = lines[j].strip()
            if not prev:
                continue
            nums = re.findall(r"[\d,]+\.\d+", prev)
            has_na = "N/A" in prev
            if len(nums) >= 3 or (has_na and len(nums) >= 2):
                # MV = 3er numero (indice 2) para caso normal QTY PRICE MV [INCOME YIELD]
                #    o ultimo numero cuando hay N/A (QTY N/A MV para Alts)
                if has_na:
                    mv = nums[-1]
                else:
                    mv = nums[2]
                positions.append(
                    {
                        "sec_id": sec_id,
                        "name": prev[:80],
                        "market_value": float(mv.replace(",", "")),
                    }
                )
                seen.add(sec_id)
                break

    # Cash
    cash = 0.0
    m = re.search(r"Global Cash Balance\s+[\d,]+\.\d{2}\s+([\d,]+\.\d{2})", full_text)
    if m:
        cash = float(m.group(1).replace(",", ""))

    # Ending Account Value
    ending_value = None
    m = re.search(r"ENDING ACCOUNT VALUE\s*\$([\d,]+\.\d{2})", full_text)
    if m:
        ending_value = float(m.group(1).replace(",", ""))

    return {"positions": positions, "cash": cash, "ending_value": ending_value}


def export_listado_values(carga_max_path: Path, out_path: Path,
                           positions_by_row: dict, cash: float, calp: float, total: float):
    """Copia la hoja Listado a un archivo separado, reemplazando las formulas
    de col D (PERCENT) con los valores numericos calculados.

    positions_by_row: {row_input_pershing: monto_usd}
    total: total del portafolio (denominador para el %)
    """
    from openpyxl.utils import get_column_letter

    wb_src = openpyxl.load_workbook(str(carga_max_path))
    ws_src = wb_src["Listado"]

    # Crear workbook nuevo con solo la hoja Listado
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = "Listado"

    # Copiar cell por cell, valores + formato
    from copy import copy
    max_row = ws_src.max_row
    max_col = ws_src.max_column
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            src_cell = ws_src.cell(row=r, column=c)
            new_cell = ws_new.cell(row=r, column=c)
            val = src_cell.value
            # Reemplazar formulas de col D (PERCENT) con valor numerico calculado
            # La formula original es ='Input Pershing'!I<row>*100
            # I<row> del Input Pershing es H<row>/H31 (monto/total)
            # Entonces PERCENT = monto/total*100 (numero plano tipo 2.28 = 2.28%)
            if c == 4 and isinstance(val, str) and val.startswith("="):
                m = re.search(r"!I(\d+)\*100", val)
                if m:
                    src_row = int(m.group(1))
                    monto = positions_by_row.get(src_row, 0)
                    pct = (monto / total * 100) if total else 0
                    new_cell.value = round(pct, 4)
                else:
                    new_cell.value = None
            else:
                new_cell.value = val
            # Copiar formato basico (no exhaustivo)
            if src_cell.has_style:
                new_cell.font = copy(src_cell.font)
                new_cell.fill = copy(src_cell.fill)
                new_cell.alignment = copy(src_cell.alignment)
                new_cell.border = copy(src_cell.border)
                new_cell.number_format = src_cell.number_format

    # Ajustar anchos de columna
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        if col_letter in ws_src.column_dimensions:
            ws_new.column_dimensions[col_letter].width = ws_src.column_dimensions[col_letter].width

    wb_new.save(str(out_path))


def backup_file(path: Path) -> Path:
    backup_dir = path.parent / ".backups"
    backup_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = backup_dir / f"{path.stem}.{ts}{path.suffix}"
    shutil.copy2(path, backup_path)
    return backup_path


def find_pdf(folder: Path) -> Path:
    """Busca el PDF Pershing mas reciente en la carpeta."""
    candidates = [p for p in folder.glob("*.pdf") if not p.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError(f"No hay PDF en {folder}")
    if len(candidates) == 1:
        return candidates[0]
    # Elegir el mas reciente por mtime
    return max(candidates, key=lambda p: p.stat().st_mtime)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf", type=str, default=None,
                    help="Path del PDF Pershing (default: busca en carpeta)")
    ap.add_argument("--excel", type=str, default=str(DEFAULT_EXCEL),
                    help="Path del Excel Carga Max (default: Dropbox/BIG/2026/Export %% Maximus/)")
    ap.add_argument("--calp", type=float, default=DEFAULT_CALP,
                    help=f"Monto CALP (Carlyle AlpInvest) hardcoded (default: ${DEFAULT_CALP:,.0f})")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--tolerance", type=float, default=1.0,
                    help="Tolerancia (USD) para validacion total vs Ending Account Value")
    args = ap.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise FileNotFoundError(f"No encuentro Excel: {excel_path}")

    pdf_path = Path(args.pdf) if args.pdf else find_pdf(excel_path.parent)
    if not pdf_path.exists():
        raise FileNotFoundError(f"No encuentro PDF: {pdf_path}")

    print(f"[{datetime.now().isoformat(timespec='seconds')}] Carga Maximus")
    print(f"  PDF:   {pdf_path.name}")
    print(f"  Excel: {excel_path.name}")
    print(f"  CALP:  ${args.calp:,.2f}")
    print()

    # 1) Parse PDF
    parsed = parse_pershing_pdf(pdf_path)
    positions = parsed["positions"]
    cash = parsed["cash"]
    ending_value = parsed["ending_value"]
    total_from_pdf = sum(p["market_value"] for p in positions) + cash

    print(f"  Posiciones detectadas: {len(positions)}")
    print(f"  Cash Global Balance:   ${cash:,.2f}")
    print(f"  Total sumado:          ${total_from_pdf:,.2f}")
    if ending_value:
        diff = abs(total_from_pdf - ending_value)
        status = "OK" if diff < args.tolerance else f"MISMATCH (diff ${diff:,.2f})"
        print(f"  Ending Account Value:  ${ending_value:,.2f}  [{status}]")
        if diff >= args.tolerance:
            print(f"\n  ERROR: total del PDF no matchea el Ending Account Value (tolerance ${args.tolerance})")
            return
    print()

    # 2) Chequear que todas las positions tengan mapping
    unmapped = [p for p in positions if p["sec_id"] not in SEC_ID_TO_ROW]
    if unmapped:
        print(f"  ERROR: {len(unmapped)} posicion(es) con SecID no mapeado:")
        for p in unmapped:
            print(f"    - {p['sec_id']}: {p['name']}  (${p['market_value']:,.2f})")
        print(f"  Agregar el mapping en SEC_ID_TO_ROW dentro del script.")
        return

    mapped_ids = {p["sec_id"] for p in positions}
    missing_ids = set(SEC_ID_TO_ROW.keys()) - mapped_ids
    if missing_ids:
        print(f"  WARN: {len(missing_ids)} SecID esperado(s) NO en el PDF:")
        for sid in missing_ids:
            print(f"    - {sid} (row {SEC_ID_TO_ROW[sid]})")
        print(f"  Row(s) van a quedar con monto = 0.")

    # 3) Backup + abrir Excel
    if not args.dry_run:
        bpath = backup_file(excel_path)
        print(f"  Backup: {bpath.name}")

    wb = openpyxl.load_workbook(str(excel_path))
    ws = wb["Input Pershing"]

    # 4) Escribir montos
    # Primero limpiar rows 6-30 col H (para que missing_ids queden con 0)
    for row in range(CASH_ROW, CALP_ROW + 1):
        ws.cell(row=row, column=8).value = 0.0

    # Cash
    ws.cell(row=CASH_ROW, column=8).value = cash
    # Posiciones
    for p in positions:
        row = SEC_ID_TO_ROW[p["sec_id"]]
        ws.cell(row=row, column=8).value = p["market_value"]
    # CALP manual
    ws.cell(row=CALP_ROW, column=8).value = args.calp

    # Total previsto
    total_final = sum(p["market_value"] for p in positions) + cash + args.calp

    print()
    print(f"  === Resumen ===")
    print(f"  Total Pershing (statement): ${total_from_pdf:,.2f}")
    print(f"  + CALP (manual):            ${args.calp:,.2f}")
    print(f"  = Total Portfolio Maximus:  ${total_final:,.2f}")

    if args.dry_run:
        print(f"\n  DRY RUN - no se guarda")
        return

    # 5) Guardar
    try:
        wb.save(str(excel_path))
    except PermissionError:
        print(f"\n  ERROR: Excel abierto o lockeado. Cerra el archivo y volve a correr.")
        return
    print(f"  Guardado: {excel_path.name}")

    # 6) Reporte final + Export Listado con VALORES
    print()
    print(f"  === Composicion final Maximus ===")

    # Armar dict positions_by_row para el export
    positions_by_row = {CASH_ROW: cash, CALP_ROW: args.calp}
    for p in positions:
        positions_by_row[SEC_ID_TO_ROW[p["sec_id"]]] = p["market_value"]

    # Export Listado con valores calculados (no formulas) para subir a Maximus
    # Nombre del archivo con mes/anio del PDF
    pdf_stem = pdf_path.stem  # "Junio Big Statement" -> lo uso como sufijo
    out_listado = excel_path.parent / f"Listado_Maximus_{pdf_stem}.xlsx"
    try:
        export_listado_values(
            excel_path, out_listado,
            positions_by_row=positions_by_row,
            cash=cash, calp=args.calp, total=total_final,
        )
        print(f"  Export Listado (valores): {out_listado.name}")
    except Exception as e:
        print(f"  WARN: fallo el export Listado: {e}")

    # Reporte composicion
    for row in range(6, CALP_ROW + 1):
        monto = positions_by_row.get(row, 0)
        pct = monto / total_final * 100 if total_final else 0
        # Nombre desde el workbook original (Input Pershing col E)
        wb_calc = openpyxl.load_workbook(str(excel_path), data_only=True)
        nombre = wb_calc["Input Pershing"].cell(row=row, column=5).value or ""
        print(f"    row {row:2d}: {str(nombre)[:40]:40s} ${monto:>13,.2f}  {pct:>5.2f}%")


if __name__ == "__main__":
    main()
