"""
portfolio_reconstructor.py
==========================
Rebuilds the BIG Fund equity sleeve timeline from Pershing transaction history.

Reads Transactions_JXD101380 History.xlsx and:
  1. Parses all BUY/SELL trades (handles multiple line formats including BRK B, parValue bonds)
  2. Builds position-over-time (qty) for each security
  3. Classifies each security by sleeve (Equity / FI / Alts / Cash)
  4. Values positions month-end using:
     - ETFs with Yahoo ticker → daily price
     - UCITS → baha monthly returns + latest NAV anchor
     - Privates → interpolated from known values
  5. Sums equity sleeve MV monthly
  6. Outputs data/equity_sleeve_real.json

Usage:
    python scripts/portfolio_reconstructor.py "C:\\path\\to\\Transactions_JXD101380 History.xlsx"
"""

import openpyxl
import re
import json
import sys
from datetime import datetime, date, timedelta
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).parent.parent

# ==============================================================
# SECURITY CLASSIFICATION
# ==============================================================
# symbol / cusip → (sleeve, ticker_pretty, isin, yahoo_ticker_for_price)
SECURITY_MAP = {
    # EQUITY
    "CSPX:GB":   ("Equity", "CSPX",  "IE00B5BMR087", "SPY"),       # SPY as proxy for S&P500
    "BRK B":     ("Equity", "BRK.B", "US0846707026", "BRK-B"),     # Berkshire Class B
    "G6430T809": ("Equity", "NBGMT", "IE00BFMHRK20", None),        # NB Megatrends (UCITS)
    "L6366L196": ("Equity", "MFSCV", "LU1985812756", None),        # MFS Contrarian
    "L468AA656": ("Equity", "JHGSC", "LU2940405447", None),        # Janus Global Small Cos
    "G5S05E528": ("Equity", "LGLI",  "IE00BF4KN675", None),        # Lazard Infra
    "G5441Y310": ("Equity", "NBRE",  "LU1648392275", None),        # NB US Real Estate (closed Feb-2026)
    "G9373W177": ("Equity", "VIRTUS","LU2020382479", None),        # Virtus US Small Cap (closed Feb-2026)
    "G8T49N214": ("Equity", "THOR",  "IE00B6YCBF59", None),        # Thornburg
    "ILF":       ("Equity", "ILF",   "US4642873909", "ILF"),
    "4BRZ:DE":   ("Equity", "4BRZ",  "DE000A0Q4R85", "EWZ"),       # use EWZ as proxy
    "ARGT":      ("Equity", "ARGT",  "US37950E2596", "ARGT"),

    # ALTS
    "GLD":       ("Alternatives", "GLD",   "US78463V1070", "GLD"),
    "IBIT":      ("Alternatives", "IBIT",  "US46438F1012", "IBIT"),
    "L6712R194": ("Alternatives", "NBPEA", "LU2659193242", None),  # NB PE Access (closed Apr-21)
    "L4R58Q111": ("Alternatives", "FLEX",  "LU2XXX",       None),  # Franklin Lex PE (new)
    "L4680C117": ("Alternatives", "HLGPIF","LU2XXX",       None),  # Hamilton Lane Infra (new)
    "449LP0055": ("Alternatives", "HLEND", "KYG4737U1085", None),  # HPS Corporate Lending
    "379LP0083": ("Alternatives", "GCRED", "GCRED-I",      None),  # Golub Capital

    # FIXED INCOME
    "G7113P361": ("Fixed Income", "PIMCO-INC", "IE00B87KCF77", None),
    "G7S11T150": ("Fixed Income", "PIMCO-LD",  "IE00BDT57R20", None),
    "G7097Y503": ("Fixed Income", "PIMCO-EM",  "IE00B29K0P99", None),
    "G5896H580": ("Fixed Income", "MANIG",     "IE000OE87WX6", None),
    "L8147L735": ("Fixed Income", "SGCB",      "LU2049315265", None),
    "L9690L577": ("Fixed Income", "WELL",      "LU-WELL",      None),  # Wellington Credit (closed)
    "G7151GAA7": ("Alternatives", "BPCC",      "XS2658535526", None),  # Barings Private Credit (parValue) — Alts en positions_latest.json
    "G5478EAA2": ("Fixed Income", "TGF",       "XS2324777171", None),  # Tenac (parValue)

    # CASH / MMF
    "L4060F300": ("Cash", "FMM1", "LU-FMM1", None),
    "L4058R662": ("Cash", "FMM2", "LU-FMM2", None),
}

# Symbols tradeados como bond parValue: qty = face value, price = per 100 of par.
# MV verdadero = qty * price / 100 (no qty * price).
PAR_VALUE_SYMS = {
    "G5478EAA2",  # TGF (Tenac)
    "G7151GAA7",  # BPCC (Barings)
}


def parse_transactions(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # Find header row
    hdr = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[0] == "Process Date":
            hdr = i
            break
    if hdr is None:
        raise ValueError("Header row not found")

    trades = []
    # Regex patterns to cover all formats in Pershing:
    # Buy 14410.58400 share(s) of G8T49N214 at 42.3300
    # Sell -2572.00000 share(s) of BRK B at 472.4350      (has space in symbol)
    # Buy 280.00000 of CSPX:GB at 707.5200                (no "share(s)")
    # Buy 40588.00000 parValue of G7151GAA7 at 123.1900  (parValue for bonds)
    # Correct Sell -178660.05000 share(s) of G7113P361 at 20.1500
    # Cancel Sell 225102.06400 share(s) of G7113P361 at 20.1500
    patterns = [
        re.compile(r"^(Buy|Sell|Correct Buy|Correct Sell)\s+(-?[\d.]+)\s+(?:share\(s\)|parValue)\s+of\s+([A-Z0-9: .]+?)\s+at\s+([\d.]+)$"),
        re.compile(r"^(Buy|Sell|Correct Buy|Correct Sell)\s+(-?[\d.]+)\s+of\s+([A-Z0-9: .]+?)\s+at\s+([\d.]+)$"),
    ]
    cancels = re.compile(r"^Cancel\s+(Buy|Sell)\s+(-?[\d.]+)\s+.*?of\s+([A-Z0-9: .]+?)\s+at\s+([\d.]+)$")

    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        date_cell, typ, desc, net_base, *_ = row
        if not date_cell or not typ:
            continue
        typ_s = str(typ).strip()

        # Skip non-trade rows (fees, dividends, etc.)
        is_trade = any(k in typ_s for k in ["Buy", "Sell"])
        if not is_trade:
            continue

        # Cancels son reversiones explicitas de un trade previo (ej: Buy ejecutado por
        # error -> Cancel Buy lo revierte, y luego un Correct Buy es el trade verdadero).
        # Parsearlos como trade opuesto para que original + cancel se compensen y solo
        # quede el Correct.
        if typ_s.startswith("Cancel"):
            m = cancels.match(typ_s)
            if not m:
                continue
            action_word, qty_raw, sym, price = m.groups()
            trade_date = date_cell if isinstance(date_cell, datetime) else None
            if not trade_date:
                continue
            # Cancel Buy -> reversa una compra: qty baja, cash vuelve -> SELL
            # Cancel Sell -> reversa una venta: qty sube, cash sale -> BUY
            side = "SELL" if action_word == "Buy" else "BUY"
            trades.append({
                "date": trade_date.date(),
                "side": side,
                "sym": sym.strip(),
                "qty": abs(float(qty_raw)),
                "price": float(price),
                "net_usd": float(net_base) if net_base else 0,
                "desc": "CANCEL: " + (desc or "")[:40],
                "raw_type": typ_s,
            })
            continue

        action = None
        qty = None
        sym = None
        price = None
        for pat in patterns:
            m = pat.match(typ_s)
            if m:
                action, qty, sym, price = m.groups()
                break
        if not action:
            continue

        trade_date = date_cell if isinstance(date_cell, datetime) else None
        if not trade_date:
            continue

        side = "SELL" if "Sell" in action else "BUY"
        qty_val = abs(float(qty))
        # Correct Sell/Buy are adjustments — treat as normal trade direction
        # But they might cancel previous trades; for simplicity treat as independent

        trades.append({
            "date": trade_date.date(),
            "side": side,
            "sym": sym.strip(),
            "qty": qty_val,
            "price": float(price),
            "net_usd": float(net_base) if net_base else 0,
            "desc": (desc or "")[:50],
            "raw_type": typ_s,
        })

    return trades


def build_positions_over_time(trades):
    """Build {sym: [{date, qty_after, avg_cost}]} timeline."""
    # Group by sym, sort by date
    by_sym = defaultdict(list)
    for t in trades:
        by_sym[t["sym"]].append(t)

    timelines = {}
    for sym, ts in by_sym.items():
        ts_sorted = sorted(ts, key=lambda x: x["date"])
        qty = 0.0
        cost_basis = 0.0
        timeline = []
        for t in ts_sorted:
            if t["side"] == "BUY":
                qty += t["qty"]
                cost_basis += t["qty"] * t["price"]
            else:  # SELL
                if qty > 0:
                    cost_basis *= max(1 - t["qty"] / qty, 0)  # proportional cost reduction
                qty -= t["qty"]
            avg_cost = cost_basis / qty if qty > 0 else None
            timeline.append({
                "date": t["date"],
                "qty_after": qty,
                "avg_cost": avg_cost,
                "trade_price": t["price"],
            })
        timelines[sym] = timeline
    return timelines


def month_end_dates(start: date, end: date):
    """Yield month-end dates between start and end (inclusive)."""
    y, m = start.year, start.month
    while True:
        # Last day of month
        if m == 12:
            next_y, next_m = y + 1, 1
        else:
            next_y, next_m = y, m + 1
        last_day = date(next_y, next_m, 1) - timedelta(days=1)
        # Cortamos cuando salimos del mes-de-end. Asi incluimos el month-end del
        # mes que contiene end (ej: si end=2026-04-21, incluye 2026-04-30).
        if (last_day.year, last_day.month) > (end.year, end.month):
            break
        if last_day >= start:
            yield last_day
        y, m = next_y, next_m


def qty_at_date(timeline, target: date):
    """Return qty held at end of target date."""
    qty = 0.0
    for t in timeline:
        if t["date"] <= target:
            qty = t["qty_after"]
        else:
            break
    return qty


def fetch_yahoo_daily(ticker: str, start: date, end: date):
    """Return {date_iso: close} via yfinance."""
    import yfinance as yf
    h = yf.Ticker(ticker).history(start=start.isoformat(), end=(end + timedelta(days=3)).isoformat())
    return {ts.strftime("%Y-%m-%d"): float(row["Close"]) for ts, row in h.iterrows()}


def load_baha_monthly_returns(isin: str):
    """Return {YYYY-MM: decimal return}."""
    fp = ROOT / "data" / "baha" / f"{isin}.json"
    if not fp.exists():
        return {}
    spanish_months = {'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,'jul':7,'ago':8,'sep':9,'oct':10,'nov':11,'dic':12}
    with open(fp) as f:
        data = json.load(f)
    rets = {}
    for year_str, months_map in data["monthlyReturns"].items():
        for m_name, pct in months_map.items():
            if m_name in spanish_months:
                rets[f"{year_str}-{spanish_months[m_name]:02d}"] = float(pct) / 100
    return rets


def build_ucits_nav_series(isin: str, anchor_nav: float, anchor_month: str):
    """Given a UCITS anchor NAV at month X, derive NAV at all other months
    using monthly returns (backward and forward)."""
    rets = load_baha_monthly_returns(isin)
    if not rets:
        return {}
    all_months = sorted(set(list(rets.keys()) + [anchor_month]))
    nav_by_month = {anchor_month: anchor_nav}
    # Forward from anchor
    cur = anchor_nav
    for m in [x for x in all_months if x > anchor_month]:
        if m in rets:
            cur *= (1 + rets[m])
            nav_by_month[m] = cur
    # Backward from anchor (inverse compounding)
    cur = anchor_nav
    for m in reversed([x for x in all_months if x < anchor_month]):
        next_m_month = m[:7]
        # rets[X] represents return DURING X; so nav at end of X-1 = nav_X / (1 + ret_X)
        # Walking backward, find return for month AFTER this one
        next_idx = all_months.index(m) + 1
        if next_idx < len(all_months):
            next_m = all_months[next_idx]
            if next_m in rets:
                cur = cur / (1 + rets[next_m])
                nav_by_month[m] = cur
    return nav_by_month


# Known current NAVs (as of April 2026) for UCITS (from Pershing positions export and baha)
# Used as anchor for NAV series reconstruction
UCITS_ANCHOR_NAV = {
    # isin → (nav, month anchor "YYYY-MM")
    "IE00BFMHRK20": (23.88, "2026-04"),   # NB Megatrends
    "LU1985812756": (290.16, "2026-04"),  # MFS Contrarian
    "LU2940405447": (54.47, "2026-04"),   # Janus Global Small Cos
    "IE00BF4KN675": (20.2315, "2026-04"), # Lazard Infra
    "IE00B6YCBF59": (42.33, "2026-04"),   # Thornburg
    # Non-equity sleeves (for completeness)
    "IE00B87KCF77": (20.22, "2026-04"),
    "IE00BDT57R20": (13.95, "2026-04"),
    "IE000OE87WX6": (124.31, "2026-04"),
    "IE00B29K0P99": (18.40, "2026-04"),
    "XS2324777171": (157.816, "2026-04"),
    "LU2049315265": (2111.85, "2026-04"),
    # Closed / illiquid positions (sale price as anchor)
    "LU1648392275": (27.88, "2026-02"),   # NB US RE sold at 27.88 in Feb
    "LU2020382479": (34.04, "2026-02"),   # Virtus US SC sold at 34.04 in Feb
    "LU2659193242": (12.2061, "2026-04"), # NB Global PE sold at 12.2061
}


def _build_sleeve_series(sleeve_name, month_ends, timelines, yahoo_cache, ucits_nav_cache,
                          pershing_override_by_ticker=None):
    """Construye serie de MV mes-a-mes para un sleeve dado.

    pershing_override_by_ticker: dict {ticker: {"value": MV_USD, "qty": qty}} del
    positions_latest.json — si se pasa, se usa SOLO para el ultimo me_date.
    Esto evita proxys Yahoo (ej. EWZ para 4BRZ) y refleja el MV real del custodio.
    """
    sleeve_series = []
    last_me = month_ends[-1] if month_ends else None
    override = pershing_override_by_ticker or {}
    for me_date in month_ends:
        total_mv = 0.0
        holdings_at_date = []
        is_last = (me_date == last_me)
        for sym, meta in SECURITY_MAP.items():
            sleeve, ticker, isin, yf_ticker = meta
            if sleeve != sleeve_name:
                continue
            if sym not in timelines:
                continue
            qty = qty_at_date(timelines[sym], me_date)
            if qty <= 0.0001:
                continue

            # Pershing override SOLO para el ultimo punto (MV USD directo del custodio).
            if is_last and ticker in override:
                ovr = override[ticker]
                mv = ovr["value"]
                # price implicito = MV / qty (para mantener formato y trazabilidad)
                ovr_qty = ovr.get("qty") or qty
                price = mv / ovr_qty if ovr_qty else 0
                total_mv += mv
                holdings_at_date.append({
                    "ticker": ticker, "qty": ovr_qty, "price": price, "mv": mv,
                    "source": "pershing_positions",
                })
                continue

            price = None
            if yf_ticker and sym in yahoo_cache:
                yc = yahoo_cache[sym]
                for d_offset in range(0, 10):
                    check_d = me_date - timedelta(days=d_offset)
                    key = check_d.isoformat()
                    if key in yc:
                        price = yc[key]
                        break
            elif sym in ucits_nav_cache:
                month_key = f"{me_date.year}-{me_date.month:02d}"
                price = ucits_nav_cache[sym].get(month_key)
            if price is None:
                # Fallback: ultimo trade price conocido
                for t in reversed(timelines[sym]):
                    if t["date"] <= me_date:
                        price = t["trade_price"]
                        break
            if price is None:
                continue
            # Bonos parValue: qty = face value, price = per 100 of par -> MV = qty * price / 100
            mv = qty * price / 100 if sym in PAR_VALUE_SYMS else qty * price
            total_mv += mv
            holdings_at_date.append({
                "ticker": ticker, "qty": qty, "price": price, "mv": mv,
            })
        sleeve_series.append({
            "date": me_date.isoformat(),
            "mv_usd": total_mv,
            "holdings": holdings_at_date,
        })
    return sleeve_series


def _compute_sleeve_output(sleeve_name, benchmark_ticker, bmk_key, output_filename,
                            trades, timelines, month_ends, yahoo_cache, ucits_nav_cache,
                            unknown, first_date, last_date,
                            pershing_override_by_ticker=None):
    """Calcula MV series + TWR + benchmark para un sleeve y escribe el JSON."""
    print(f"\n{'#' * 90}")
    print(f"#  {sleeve_name.upper()} SLEEVE")
    print(f"{'#' * 90}")

    sleeve_series = _build_sleeve_series(
        sleeve_name, month_ends, timelines, yahoo_cache, ucits_nav_cache,
        pershing_override_by_ticker=pershing_override_by_ticker,
    )

    # Print summary
    print(f"\n{'=' * 90}")
    print(f"{'Date':12s} | {'Total ' + sleeve_name + ' MV':>22s} | {'Holdings':>10s} | Composition")
    print('=' * 90)
    for pt in sleeve_series:
        comp = ", ".join(f"{h['ticker']}:${h['mv']/1000:.0f}K" for h in sorted(pt["holdings"], key=lambda x: -x["mv"])[:6])
        print(f"{pt['date']:12s} | ${pt['mv_usd']:>20,.0f} | {len(pt['holdings']):>10d} | {comp}")

    # Flows by month (filter por sleeve)
    flows_by_month = defaultdict(float)
    for t in trades:
        if t.get("sleeve") != sleeve_name:
            continue
        month_key = f"{t['date'].year}-{t['date'].month:02d}"
        flow_in = -(t["net_usd"] or 0)
        flows_by_month[month_key] += flow_in

    # TWR series
    print(f"\n{'=' * 90}")
    print(f"{sleeve_name.upper()} SLEEVE TWR CALCULATION")
    print('=' * 90)
    print(f"{'Date':12s} | {'MV End':>14s} | {'Flow In':>14s} | {'TWR %':>8s} | {'Cum Index':>10s}")
    cum_index = 100.0
    twr_series = []
    prev_mv = None
    for pt in sleeve_series:
        me_date = pt["date"]
        mv = pt["mv_usd"]
        month_key = me_date[:7]
        flow = flows_by_month.get(month_key, 0)
        twr = (mv - flow) / prev_mv - 1 if (prev_mv is not None and prev_mv > 0) else 0
        cum_index *= (1 + twr)
        print(f"{me_date:12s} | ${mv:>12,.0f} | ${flow:>12,.0f} | {twr*100:>+7.2f}% | {cum_index:>10.2f}")
        twr_series.append({"date": me_date, "mv_usd": mv, "flow_in": flow, "twr": twr, "index": round(cum_index, 4)})
        prev_mv = mv

    # Benchmark
    print(f"\n{'=' * 90}\nFETCHING {benchmark_ticker} BENCHMARK")
    # Mismo end_fetch que yahoo_cache: extender hasta hoy para que el bmk llegue
    # al today_date appended en month_ends.
    bmk_end = max(last_date, date.today()) + timedelta(days=5)
    bmk_daily = fetch_yahoo_daily(benchmark_ticker, first_date - timedelta(days=5), bmk_end)
    bmk_by_month = {}
    for me_date in month_ends:
        for d_offset in range(0, 10):
            check = (me_date - timedelta(days=d_offset)).isoformat()
            if check in bmk_daily:
                bmk_by_month[me_date.isoformat()] = bmk_daily[check]
                break

    bmk_index_series = []
    if twr_series:
        bmk_base = bmk_by_month.get(twr_series[0]["date"])
        if bmk_base:
            for pt in twr_series:
                bmk_val = bmk_by_month.get(pt["date"])
                if bmk_val:
                    bmk_index_series.append({
                        "date": pt["date"],
                        "price": bmk_val,
                        "index": round(bmk_val / bmk_base * 100, 4),
                    })

    if twr_series and bmk_index_series:
        final_sleeve = twr_series[-1]["index"]
        final_bmk = bmk_index_series[-1]["index"]
        print(f"\n{'=' * 90}")
        print(f"SI {sleeve_name.upper()} TWR:   {final_sleeve - 100:+.2f}%")
        print(f"SI {benchmark_ticker}:         {final_bmk - 100:+.2f}%")
        print(f"ALPHA SI:        {final_sleeve - final_bmk:+.2f}pp")
        status = "GANANDO" if final_sleeve > final_bmk else "PERDIENDO"
        print(f"STATUS:          {status}")

    sleeve_key = "fi" if sleeve_name == "Fixed Income" else sleeve_name.lower()
    out = {
        "refreshedAt": datetime.now().isoformat(),
        "source": "Reconstructed from Pershing Transactions export",
        "sleeve": sleeve_name,
        "benchmark": benchmark_ticker,
        "first_trade_date": first_date.isoformat(),
        "last_trade_date": last_date.isoformat(),
        "n_trades": len(trades),
        "unknown_securities": unknown,
        f"sleeve_series_{sleeve_key}": sleeve_series,
        "twr_series": twr_series,
        bmk_key: bmk_index_series,
        f"{sleeve_key}_flows_by_month": dict(flows_by_month),
        "timelines_summary": {
            sym: {
                "ticker": SECURITY_MAP[sym][1] if sym in SECURITY_MAP else "?",
                "sleeve": SECURITY_MAP[sym][0] if sym in SECURITY_MAP else "?",
                "last_qty": timelines[sym][-1]["qty_after"],
                "n_trades": len(timelines[sym]),
            } for sym in timelines
        },
    }

    out_path = ROOT / "data" / output_filename
    with open(out_path, "w") as f:
        json.dump(out, f, indent=2, default=str)
    print(f"\nSaved: {out_path}")
    return out


def main():
    if len(sys.argv) > 1:
        xlsx = Path(sys.argv[1])
    else:
        xlsx = Path("C:/Users/lmonp/Downloads/Transactions_JXD101380 History.xlsx")
    print(f"[{datetime.now()}] Parsing: {xlsx.name}")

    trades = parse_transactions(xlsx)
    print(f"  Parsed {len(trades)} trades")

    all_syms = set(t["sym"] for t in trades)
    print(f"  Unique securities: {len(all_syms)}")
    unknown = [s for s in all_syms if s not in SECURITY_MAP]
    if unknown:
        print(f"  WARNING — unknown securities: {unknown}")

    trades_by_sleeve = defaultdict(list)
    for t in trades:
        meta = SECURITY_MAP.get(t["sym"])
        if not meta:
            continue
        t["sleeve"] = meta[0]
        t["ticker"] = meta[1]
        t["isin"] = meta[2]
        t["yahoo"] = meta[3]
        trades_by_sleeve[meta[0]].append(t)
    for sl in ["Equity", "Alternatives", "Fixed Income", "Cash"]:
        print(f"  {sl}: {len(trades_by_sleeve[sl])} trades")

    timelines = build_positions_over_time(trades)
    first_date = min(t["date"] for t in trades)
    last_date = max(t["date"] for t in trades)
    # Extender hasta hoy para que ACWI/AGG benchmarks y precios live lleguen al
    # mismo punto que today_date (sino el cuadro Multi-Period queda con bench stale).
    end_fetch = max(last_date, date.today()) + timedelta(days=5)
    print(f"\n  Period: {first_date} to {last_date}  (fetch hasta {end_fetch})")

    # Pre-fetch caches para Equity + Fixed Income
    SLEEVES_OF_INTEREST = {"Equity", "Fixed Income"}

    yahoo_cache = {}
    for sym, meta in SECURITY_MAP.items():
        sleeve, ticker, isin, yf_ticker = meta
        if sleeve not in SLEEVES_OF_INTEREST or not yf_ticker:
            continue
        print(f"  Fetching {yf_ticker} (for {ticker})...")
        yahoo_cache[sym] = fetch_yahoo_daily(yf_ticker, first_date - timedelta(days=10), end_fetch)

    ucits_nav_cache = {}
    for sym, meta in SECURITY_MAP.items():
        sleeve, ticker, isin, yf_ticker = meta
        if sleeve not in SLEEVES_OF_INTEREST or yf_ticker:
            continue
        if isin in UCITS_ANCHOR_NAV:
            nav, anchor_m = UCITS_ANCHOR_NAV[isin]
            ucits_nav_cache[sym] = build_ucits_nav_series(isin, nav, anchor_m)
            print(f"  UCITS NAV series for {ticker}: {len(ucits_nav_cache[sym])} months")

    month_ends = list(month_end_dates(first_date, last_date))
    today_date = date.today()
    if month_ends and today_date > month_ends[-1]:
        month_ends.append(today_date)

    # Cargar positions_latest.json para override del ultimo punto (MV USD del custodio).
    # Evita proxys Yahoo (ej. EWZ para 4BRZ) y refleja la valuacion real Pershing.
    pershing_override = {}
    positions_path = ROOT / "data" / "positions_latest.json"
    if positions_path.exists():
        try:
            pl = json.load(open(positions_path))
            for pos in pl.get("positions", []):
                t = pos.get("ticker")
                if t:
                    pershing_override[t] = {"value": pos.get("value"), "qty": pos.get("qty")}
            print(f"\n  Pershing override loaded: {len(pershing_override)} positions @ {pl.get('as_of','?')}")
        except Exception as e:
            print(f"\n  Pershing override skipped: {e}")

    # Equity vs ACWI
    _compute_sleeve_output(
        sleeve_name="Equity",
        benchmark_ticker="ACWI",
        bmk_key="acwi_index_series",
        output_filename="equity_sleeve_real.json",
        trades=trades, timelines=timelines, month_ends=month_ends,
        yahoo_cache=yahoo_cache, ucits_nav_cache=ucits_nav_cache,
        unknown=unknown, first_date=first_date, last_date=last_date,
        pershing_override_by_ticker=pershing_override,
    )

    # Fixed Income vs AGG
    _compute_sleeve_output(
        sleeve_name="Fixed Income",
        benchmark_ticker="AGG",
        bmk_key="agg_index_series",
        output_filename="fi_sleeve_real.json",
        trades=trades, timelines=timelines, month_ends=month_ends,
        yahoo_cache=yahoo_cache, ucits_nav_cache=ucits_nav_cache,
        unknown=unknown, first_date=first_date, last_date=last_date,
        pershing_override_by_ticker=pershing_override,
    )


if __name__ == "__main__":
    main()
