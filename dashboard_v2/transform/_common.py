"""
Helpers compartidos por los parsers.

Los XLSX de NetX360+ tienen todos el mismo layout:
- Rows 1-N: metadata (Account, Client, Duration, As of...)
- Row N+1: header con nombres de columnas
- Rows N+2..: data

Este modulo abstrae la lectura del layout comun.
"""
from __future__ import annotations

from datetime import datetime, date
from pathlib import Path

import openpyxl


# Repo root — dinámico (Windows local + Linux CI). Este archivo está en
# dashboard_v2/transform/_common.py → parents[2] = repo root.
ROOT = Path(__file__).resolve().parents[2]


def find_header_row(ws, min_filled_cells: int = 5, max_scan: int = 30) -> int:
    """Encuentra la row del header: primera row con >=min_filled_cells consecutivas."""
    for r in range(1, min(ws.max_row + 1, max_scan)):
        filled = sum(
            1 for c in range(1, ws.max_column + 1) if ws.cell(row=r, column=c).value is not None
        )
        if filled >= min_filled_cells:
            return r
    raise ValueError(f"No header row found (scanned {max_scan} rows)")


def read_metadata(ws, header_row: int) -> dict:
    """Extrae metadata de las rows arriba del header (formato 'Key: Value')."""
    meta = {}
    for r in range(1, header_row):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        s = str(v).strip()
        if ":" in s:
            k, val = s.split(":", 1)
            meta[k.strip().lower()] = val.strip()
        else:
            # Row sin ":" — es el titulo (ej: "Positions")
            meta.setdefault("_title", s)
    return meta


def parse_header_and_rows(path: Path) -> tuple[dict, list[str], list[dict]]:
    """
    Retorna: (metadata, column_names, data_rows_as_dicts).

    Los data rows son dicts {column_name: value} donde value tiene el tipo Python
    nativo de openpyxl (str, int, float, datetime).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = find_header_row(ws)
    metadata = read_metadata(ws, header_row)

    # Column names
    columns = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        columns.append(str(v).strip() if v is not None else f"_col{c}")

    # Data rows
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {}
        empty = True
        for c, col in enumerate(columns, start=1):
            v = ws.cell(row=r, column=c).value
            row[col] = v
            if v is not None:
                empty = False
        if not empty:
            rows.append(row)

    return metadata, columns, rows


def to_float(v, default=None) -> float | None:
    """Convierte v a float. Retorna default si es None, '-', ''."""
    if v is None or v == "" or v == "-":
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def to_str(v, default=None) -> str | None:
    """Convierte v a str. Retorna default si es None o vacio o '-'."""
    if v is None:
        return default
    s = str(v).strip()
    if s in ("", "-", "N/A"):
        return default
    return s


def to_iso_date(v, default=None) -> str | None:
    """Convierte v a YYYY-MM-DD. Acepta datetime, date, o strings comunes."""
    if v is None or v == "" or v == "-":
        return default
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if s in ("", "-", "N/A", "Multiple"):
        return default
    # Probar formatos comunes NetX360+
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%b %d, %Y %I:%M %p EDT",
        "%b %d, %Y %H:%M:%S EDT",
        "%b %d, %Y %I:%M:%S %p EDT",
        "%b %d, %Y",
    ):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return default


def utc_now_iso() -> str:
    """Timestamp UTC ISO 8601 sin microseconds."""
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def relpath_from_root(p: Path) -> str:
    """Path relativo a ROOT como str POSIX-style (para JSON portable)."""
    p = Path(p).resolve()
    try:
        return p.relative_to(ROOT).as_posix()
    except ValueError:
        return p.as_posix()
